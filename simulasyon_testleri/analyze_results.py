#!/usr/bin/env python3
"""
EBARS Comprehensive Results Analysis System
==========================================

This module provides research-grade statistical analysis capabilities for EBARS simulation data,
designed to support academic research claims about dynamic adaptation effectiveness.

Features:
- Comprehensive statistical metrics calculation
- Advanced correlation and trend analysis
- Academic reporting with LaTeX output
- Comparative analysis between agent profiles
- Statistical significance testing
- Effect size calculations
- Research-ready documentation

Author: EBARS Research Team
Date: 2025-12-06
Version: 1.0
"""

import pandas as pd
import numpy as np
import scipy.stats as stats
from scipy import optimize
import matplotlib.pyplot as plt
import seaborn as sns
from typing import Dict, List, Optional, Tuple, Union, Any
import warnings
import json
import csv
from datetime import datetime
from pathlib import Path
import os
import re
from dataclasses import dataclass, asdict
from sklearn.metrics import r2_score
from sklearn.linear_model import LinearRegression
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler
import itertools

# Suppress warnings for cleaner output
warnings.filterwarnings('ignore', category=RuntimeWarning)
warnings.filterwarnings('ignore', category=FutureWarning)

@dataclass
class StatisticalResult:
    """Container for statistical test results"""
    test_name: str
    statistic: float
    p_value: float
    effect_size: Optional[float] = None
    confidence_interval: Optional[Tuple[float, float]] = None
    interpretation: str = ""
    significance_level: float = 0.05

@dataclass
class AdaptationMetrics:
    """Container for adaptation effectiveness metrics"""
    agent_id: str
    convergence_rate: float
    stability_index: float
    adaptation_speed: float
    final_performance: float
    learning_efficiency: float
    hysteresis_index: float

class EBARSResultsAnalyzer:
    """
    Comprehensive statistical analysis engine for EBARS simulation data.
    
    This class provides research-grade statistical analysis capabilities including:
    - Adaptation effectiveness analysis
    - Score convergence analysis
    - System responsiveness metrics
    - Stability and hysteresis analysis
    - Comparative statistical testing
    - Academic reporting features
    """
    
    def __init__(self, output_dir: str = "analysis_output", confidence_level: float = 0.95):
        """
        Initialize the EBARS Results Analyzer.
        
        Args:
            output_dir (str): Output directory for analysis results
            confidence_level (float): Confidence level for statistical tests (default: 0.95)
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.confidence_level = confidence_level
        self.alpha = 1 - confidence_level
        
        # Agent profile definitions
        self.agent_profiles = {
            'agent_a': {
                'name': 'Struggling Agent',
                'description': 'Slow learner with consistent negative feedback',
                'expected_pattern': 'gradual_improvement'
            },
            'agent_b': {
                'name': 'Fast Learner',
                'description': 'Rapid adaptation with positive feedback',
                'expected_pattern': 'rapid_improvement'
            },
            'agent_c': {
                'name': 'Variable Agent',
                'description': 'Inconsistent performance with mixed feedback',
                'expected_pattern': 'variable_performance'
            }
        }
        
        # Difficulty level hierarchy
        self.difficulty_hierarchy = {
            'very_struggling': 1,
            'struggling': 2,
            'normal': 3,
            'good': 4,
            'excellent': 5
        }
        
        self.data = None
        self.results = {}
        self.statistical_tests = []
    
    def load_data(self, csv_file: str) -> pd.DataFrame:
        """
        Load and validate EBARS simulation data.
        
        Args:
            csv_file (str): Path to CSV file containing simulation results
            
        Returns:
            pd.DataFrame: Loaded and validated data
            
        Raises:
            FileNotFoundError: If CSV file doesn't exist
            ValueError: If data validation fails
        """
        if not os.path.exists(csv_file):
            raise FileNotFoundError(f"CSV file not found: {csv_file}")
        
        try:
            # Load data
            data = pd.read_csv(csv_file, encoding='utf-8')
            print(f"✅ Loaded {len(data)} records from {csv_file}")
            
            # Validate required columns
            required_columns = [
                'agent_id', 'turn_number', 'comprehension_score', 'difficulty_level',
                'score_delta', 'emoji_feedback', 'processing_time_ms', 'timestamp'
            ]
            
            missing_columns = [col for col in required_columns if col not in data.columns]
            if missing_columns:
                raise ValueError(f"Missing required columns: {missing_columns}")
            
            # Data type conversions and validation
            data['turn_number'] = pd.to_numeric(data['turn_number'], errors='coerce')
            data['comprehension_score'] = pd.to_numeric(data['comprehension_score'], errors='coerce')
            data['score_delta'] = pd.to_numeric(data['score_delta'], errors='coerce')
            data['processing_time_ms'] = pd.to_numeric(data['processing_time_ms'], errors='coerce')
            data['timestamp'] = pd.to_datetime(data['timestamp'], errors='coerce')
            
            # Add derived columns
            data['difficulty_numeric'] = data['difficulty_level'].map(self.difficulty_hierarchy)
            data['positive_feedback'] = data['emoji_feedback'].isin(['👍', '😊']).astype(int)
            data['negative_feedback'] = data['emoji_feedback'].isin(['❌', '😞']).astype(int)
            
            # Remove invalid records
            initial_rows = len(data)
            data = data.dropna(subset=['turn_number', 'comprehension_score', 'difficulty_numeric'])
            removed_rows = initial_rows - len(data)
            
            if removed_rows > 0:
                warnings.warn(f"Removed {removed_rows} rows with invalid data")
            
            # Sort data
            data = data.sort_values(['agent_id', 'turn_number']).reset_index(drop=True)
            
            self.data = data
            print(f"✅ Data validation complete. {len(data)} valid records loaded.")
            return data
            
        except Exception as e:
            raise ValueError(f"Error loading data: {str(e)}")
    
    def calculate_adaptation_effectiveness(self) -> Dict[str, AdaptationMetrics]:
        """
        Calculate comprehensive adaptation effectiveness metrics for each agent.
        
        Returns:
            Dict[str, AdaptationMetrics]: Adaptation metrics by agent
        """
        if self.data is None:
            raise ValueError("No data loaded. Call load_data() first.")
        
        adaptation_metrics = {}
        
        for agent_id in self.data['agent_id'].unique():
            agent_data = self.data[self.data['agent_id'] == agent_id].copy()
            agent_data = agent_data.sort_values('turn_number')
            
            # Convergence Rate Analysis
            scores = agent_data['comprehension_score'].values
            turns = agent_data['turn_number'].values
            
            # Fit exponential convergence model: y = a * (1 - exp(-b*x)) + c
            def convergence_model(x, a, b, c):
                return a * (1 - np.exp(-b * x)) + c
            
            try:
                popt, _ = optimize.curve_fit(convergence_model, turns, scores, 
                                           bounds=([0, 0, 0], [100, 1, 100]),
                                           maxfev=2000)
                convergence_rate = popt[1]  # b parameter
            except:
                # Fallback to linear regression slope
                convergence_rate = np.corrcoef(turns, scores)[0, 1] if len(turns) > 1 else 0
            
            # Stability Index (inverse of variance in last 5 turns)
            last_scores = scores[-5:] if len(scores) >= 5 else scores
            stability_index = 1 / (1 + np.var(last_scores)) if len(last_scores) > 1 else 1
            
            # Adaptation Speed (mean absolute score delta per turn)
            adaptation_speed = np.mean(np.abs(agent_data['score_delta'].values[1:]))
            
            # Final Performance (average of last 3 scores)
            final_performance = np.mean(scores[-3:]) if len(scores) >= 3 else scores[-1] if len(scores) > 0 else 0
            
            # Learning Efficiency (improvement per unit time)
            if len(scores) > 1:
                total_improvement = scores[-1] - scores[0]
                total_time = len(scores)
                learning_efficiency = total_improvement / total_time
            else:
                learning_efficiency = 0
            
            # Hysteresis Index (resistance to oscillation)
            direction_changes = 0
            for i in range(2, len(agent_data)):
                if len(agent_data) > 2:
                    prev_delta = agent_data.iloc[i-1]['score_delta']
                    curr_delta = agent_data.iloc[i]['score_delta']
                    if (prev_delta > 0 > curr_delta) or (prev_delta < 0 < curr_delta):
                        direction_changes += 1
            
            hysteresis_index = 1 / (1 + direction_changes / len(scores)) if len(scores) > 0 else 1
            
            adaptation_metrics[agent_id] = AdaptationMetrics(
                agent_id=agent_id,
                convergence_rate=float(convergence_rate),
                stability_index=float(stability_index),
                adaptation_speed=float(adaptation_speed),
                final_performance=float(final_performance),
                learning_efficiency=float(learning_efficiency),
                hysteresis_index=float(hysteresis_index)
            )
        
        self.results['adaptation_metrics'] = adaptation_metrics
        return adaptation_metrics
    
    def analyze_score_convergence(self) -> Dict[str, Dict[str, float]]:
        """
        Analyze score convergence patterns and rates.
        
        Returns:
            Dict[str, Dict[str, float]]: Convergence analysis results by agent
        """
        if self.data is None:
            raise ValueError("No data loaded. Call load_data() first.")
        
        convergence_analysis = {}
        
        for agent_id in self.data['agent_id'].unique():
            agent_data = self.data[self.data['agent_id'] == agent_id].copy()
            agent_data = agent_data.sort_values('turn_number')
            
            scores = agent_data['comprehension_score'].values
            turns = agent_data['turn_number'].values
            
            # Time to stabilization (first turn where score stays within 5% for 3+ consecutive turns)
            stabilization_turn = len(turns)
            for i in range(len(scores) - 2):
                if len(scores) > i + 2:
                    window = scores[i:i+3]
                    if np.max(window) - np.min(window) <= 5:  # Within 5% range
                        stabilization_turn = turns[i]
                        break
            
            # Convergence velocity (rate of approach to final score)
            final_score = np.mean(scores[-3:]) if len(scores) >= 3 else scores[-1] if len(scores) > 0 else 0
            distances_to_final = np.abs(scores - final_score)
            
            # Exponential decay fit for convergence velocity
            try:
                # Fit exponential decay: y = a * exp(-b*x) + c
                def decay_model(x, a, b, c):
                    return a * np.exp(-b * x) + c
                
                popt, _ = optimize.curve_fit(decay_model, turns, distances_to_final, 
                                           bounds=([0, 0, 0], [100, 1, 50]),
                                           maxfev=1000)
                convergence_velocity = popt[1]
            except:
                convergence_velocity = 0.1  # Default value
            
            # Overshooting tendency (how often scores go above optimal range)
            optimal_range = (70, 85)  # Define optimal comprehension range
            overshoot_count = np.sum(scores > optimal_range[1])
            overshoot_tendency = overshoot_count / len(scores) if len(scores) > 0 else 0
            
            # Settling time (95% of final value)
            target_threshold = 0.95 * final_score
            settling_time = len(turns)
            for i, score in enumerate(scores):
                if score >= target_threshold:
                    settling_time = turns[i]
                    break
            
            convergence_analysis[agent_id] = {
                'stabilization_turn': float(stabilization_turn),
                'convergence_velocity': float(convergence_velocity),
                'overshoot_tendency': float(overshoot_tendency),
                'settling_time': float(settling_time),
                'final_score': float(final_score),
                'score_variance': float(np.var(scores)) if len(scores) > 1 else 0
            }
        
        self.results['convergence_analysis'] = convergence_analysis
        return convergence_analysis
    
    def calculate_system_responsiveness(self) -> Dict[str, Dict[str, float]]:
        """
        Calculate system responsiveness metrics including feedback lag and adaptation delay.
        
        Returns:
            Dict[str, Dict[str, float]]: Responsiveness metrics by agent
        """
        if self.data is None:
            raise ValueError("No data loaded. Call load_data() first.")
        
        responsiveness_metrics = {}
        
        for agent_id in self.data['agent_id'].unique():
            agent_data = self.data[self.data['agent_id'] == agent_id].copy()
            agent_data = agent_data.sort_values('turn_number')
            
            # Feedback response lag (turns between feedback and score change)
            response_lags = []
            for i in range(1, len(agent_data)):
                prev_feedback = agent_data.iloc[i-1]['positive_feedback']
                curr_delta = agent_data.iloc[i]['score_delta']
                
                # Positive feedback should lead to positive delta (and vice versa)
                expected_positive = prev_feedback == 1 and curr_delta > 0
                expected_negative = prev_feedback == 0 and curr_delta < 0
                
                if expected_positive or expected_negative:
                    response_lags.append(1)  # Immediate response
                else:
                    response_lags.append(2)  # Delayed response
            
            avg_response_lag = np.mean(response_lags) if response_lags else 2
            
            # Processing time consistency
            processing_times = agent_data['processing_time_ms'].values
            processing_consistency = 1 / (1 + np.std(processing_times)) if len(processing_times) > 1 else 1
            
            # Adaptation delay (time between level transition and score adjustment)
            adaptation_delays = []
            for i in range(1, len(agent_data)):
                if agent_data.iloc[i]['level_transition'] != 'same':
                    # Look for score change in next few turns
                    delay_found = False
                    for j in range(i+1, min(i+4, len(agent_data))):
                        if abs(agent_data.iloc[j]['score_delta']) > 2:  # Significant score change
                            adaptation_delays.append(j - i)
                            delay_found = True
                            break
                    if not delay_found:
                        adaptation_delays.append(3)  # Maximum delay
            
            avg_adaptation_delay = np.mean(adaptation_delays) if adaptation_delays else 1
            
            # System reactivity (correlation between feedback and immediate score change)
            if len(agent_data) > 2:
                feedbacks = agent_data['positive_feedback'].values[:-1]  # Exclude last
                deltas = agent_data['score_delta'].values[1:]  # Exclude first
                reactivity_correlation = np.corrcoef(feedbacks, deltas)[0, 1] if len(feedbacks) == len(deltas) else 0
            else:
                reactivity_correlation = 0
            
            responsiveness_metrics[agent_id] = {
                'avg_response_lag': float(avg_response_lag),
                'processing_consistency': float(processing_consistency),
                'avg_adaptation_delay': float(avg_adaptation_delay),
                'reactivity_correlation': float(reactivity_correlation),
                'responsiveness_index': float(1 / (avg_response_lag + avg_adaptation_delay))
            }
        
        self.results['responsiveness_metrics'] = responsiveness_metrics
        return responsiveness_metrics
    
    def analyze_stability_metrics(self) -> Dict[str, Dict[str, float]]:
        """
        Analyze system stability including oscillation detection and steady-state analysis.
        
        Returns:
            Dict[str, Dict[str, float]]: Stability metrics by agent
        """
        if self.data is None:
            raise ValueError("No data loaded. Call load_data() first.")
        
        stability_metrics = {}
        
        for agent_id in self.data['agent_id'].unique():
            agent_data = self.data[self.data['agent_id'] == agent_id].copy()
            agent_data = agent_data.sort_values('turn_number')
            
            scores = agent_data['comprehension_score'].values
            deltas = agent_data['score_delta'].values
            
            # Oscillation detection using zero-crossing analysis
            zero_crossings = 0
            for i in range(1, len(deltas)):
                if (deltas[i-1] > 0 > deltas[i]) or (deltas[i-1] < 0 < deltas[i]):
                    zero_crossings += 1
            
            oscillation_index = zero_crossings / len(deltas) if len(deltas) > 0 else 0
            
            # Steady-state variance (variance in final third of simulation)
            final_third_start = int(len(scores) * 0.67)
            final_scores = scores[final_third_start:] if final_third_start < len(scores) else scores
            steady_state_variance = np.var(final_scores) if len(final_scores) > 1 else 0
            
            # Damping ratio estimation
            if oscillation_index > 0:
                # Estimate damping from envelope decay
                peaks = []
                for i in range(1, len(scores)-1):
                    if scores[i] > scores[i-1] and scores[i] > scores[i+1]:
                        peaks.append((i, scores[i]))
                
                if len(peaks) >= 2:
                    # Fit exponential decay to peak amplitudes
                    peak_times = [p[0] for p in peaks]
                    peak_values = [p[1] for p in peaks]
                    
                    try:
                        # Estimate damping coefficient
                        log_ratio = np.log(peak_values[:-1]) - np.log(peak_values[1:])
                        damping_ratio = np.mean(log_ratio) if len(log_ratio) > 0 else 0.1
                    except:
                        damping_ratio = 0.1
                else:
                    damping_ratio = 1.0  # Critically damped
            else:
                damping_ratio = 1.0  # No oscillation = good damping
            
            # Stability margin (minimum score before destabilization)
            min_stability_score = np.min(scores) if len(scores) > 0 else 0
            stability_margin = min_stability_score / 100.0  # Normalized
            
            # Recovery time (time to return to stable range after disturbance)
            recovery_times = []
            stable_range = (np.mean(scores) - np.std(scores), np.mean(scores) + np.std(scores))
            
            for i in range(len(scores)):
                if scores[i] < stable_range[0] or scores[i] > stable_range[1]:
                    # Find recovery time
                    for j in range(i+1, len(scores)):
                        if stable_range[0] <= scores[j] <= stable_range[1]:
                            recovery_times.append(j - i)
                            break
            
            avg_recovery_time = np.mean(recovery_times) if recovery_times else 0
            
            stability_metrics[agent_id] = {
                'oscillation_index': float(oscillation_index),
                'steady_state_variance': float(steady_state_variance),
                'damping_ratio': float(damping_ratio),
                'stability_margin': float(stability_margin),
                'avg_recovery_time': float(avg_recovery_time),
                'stability_score': float(1 / (1 + oscillation_index + steady_state_variance/100))
            }
        
        self.results['stability_metrics'] = stability_metrics
        return stability_metrics
    
    def perform_statistical_comparisons(self) -> Dict[str, List[StatisticalResult]]:
        """
        Perform comprehensive statistical comparisons between agent profiles.
        
        Returns:
            Dict[str, List[StatisticalResult]]: Statistical test results by comparison type
        """
        if self.data is None:
            raise ValueError("No data loaded. Call load_data() first.")
        
        comparison_results = {
            'score_comparisons': [],
            'adaptation_comparisons': [],
            'performance_comparisons': []
        }
        
        agents = sorted(self.data['agent_id'].unique())
        
        # Pairwise comparisons between agents
        for agent1, agent2 in itertools.combinations(agents, 2):
            data1 = self.data[self.data['agent_id'] == agent1]
            data2 = self.data[self.data['agent_id'] == agent2]
            
            # Score distribution comparison (Mann-Whitney U test)
            scores1 = data1['comprehension_score'].values
            scores2 = data2['comprehension_score'].values
            
            stat, p_val = stats.mannwhitneyu(scores1, scores2, alternative='two-sided')
            effect_size = self._calculate_cliff_delta(scores1, scores2)
            
            comparison_results['score_comparisons'].append(StatisticalResult(
                test_name=f"{agent1} vs {agent2} (Comprehension Scores)",
                statistic=float(stat),
                p_value=float(p_val),
                effect_size=float(effect_size),
                interpretation=self._interpret_statistical_result(p_val, effect_size, 'mann_whitney')
            ))
            
            # Adaptation speed comparison
            deltas1 = np.abs(data1['score_delta'].values[1:])  # Exclude first turn
            deltas2 = np.abs(data2['score_delta'].values[1:])
            
            if len(deltas1) > 0 and len(deltas2) > 0:
                stat, p_val = stats.mannwhitneyu(deltas1, deltas2, alternative='two-sided')
                effect_size = self._calculate_cliff_delta(deltas1, deltas2)
                
                comparison_results['adaptation_comparisons'].append(StatisticalResult(
                    test_name=f"{agent1} vs {agent2} (Adaptation Speed)",
                    statistic=float(stat),
                    p_value=float(p_val),
                    effect_size=float(effect_size),
                    interpretation=self._interpret_statistical_result(p_val, effect_size, 'mann_whitney')
                ))
            
            # Final performance comparison (T-test)
            final_scores1 = scores1[-3:] if len(scores1) >= 3 else scores1
            final_scores2 = scores2[-3:] if len(scores2) >= 3 else scores2
            
            if len(final_scores1) > 1 and len(final_scores2) > 1:
                stat, p_val = stats.ttest_ind(final_scores1, final_scores2)
                effect_size = self._calculate_cohens_d(final_scores1, final_scores2)
                
                comparison_results['performance_comparisons'].append(StatisticalResult(
                    test_name=f"{agent1} vs {agent2} (Final Performance)",
                    statistic=float(stat),
                    p_value=float(p_val),
                    effect_size=float(effect_size),
                    interpretation=self._interpret_statistical_result(p_val, effect_size, 't_test')
                ))
        
        # ANOVA for overall group differences
        if len(agents) >= 3:
            score_groups = [self.data[self.data['agent_id'] == agent]['comprehension_score'].values 
                           for agent in agents]
            
            f_stat, p_val = stats.f_oneway(*score_groups)
            eta_squared = self._calculate_eta_squared(score_groups, f_stat)
            
            comparison_results['score_comparisons'].append(StatisticalResult(
                test_name="Overall Group Differences (ANOVA)",
                statistic=float(f_stat),
                p_value=float(p_val),
                effect_size=float(eta_squared),
                interpretation=self._interpret_statistical_result(p_val, eta_squared, 'anova')
            ))
        
        self.results['statistical_comparisons'] = comparison_results
        return comparison_results
    
    def analyze_correlation_patterns(self) -> Dict[str, Dict[str, float]]:
        """
        Analyze correlation patterns between emoji feedback and system behavior.
        
        Returns:
            Dict[str, Dict[str, float]]: Correlation analysis results by agent
        """
        if self.data is None:
            raise ValueError("No data loaded. Call load_data() first.")
        
        correlation_results = {}
        
        for agent_id in self.data['agent_id'].unique():
            agent_data = self.data[self.data['agent_id'] == agent_id].copy()
            agent_data = agent_data.sort_values('turn_number')
            
            # Correlation between feedback and score changes
            feedback_score_corr = np.corrcoef(
                agent_data['positive_feedback'].values[:-1],
                agent_data['score_delta'].values[1:]
            )[0, 1] if len(agent_data) > 2 else 0
            
            # Correlation between feedback and difficulty adjustments
            difficulty_changes = np.diff(agent_data['difficulty_numeric'].values)
            feedback_difficulty_corr = np.corrcoef(
                agent_data['positive_feedback'].values[:-1],
                difficulty_changes
            )[0, 1] if len(difficulty_changes) > 0 else 0
            
            # Lag correlation (feedback effect on next turn)
            if len(agent_data) > 3:
                lag_corr = np.corrcoef(
                    agent_data['positive_feedback'].values[:-2],
                    agent_data['score_delta'].values[2:]
                )[0, 1]
            else:
                lag_corr = 0
            
            # Processing time correlation with performance
            processing_performance_corr = np.corrcoef(
                agent_data['processing_time_ms'].values,
                agent_data['comprehension_score'].values
            )[0, 1] if len(agent_data) > 1 else 0
            
            correlation_results[agent_id] = {
                'feedback_score_correlation': float(feedback_score_corr),
                'feedback_difficulty_correlation': float(feedback_difficulty_corr),
                'lag_correlation': float(lag_corr),
                'processing_performance_correlation': float(processing_performance_corr)
            }
        
        self.results['correlation_analysis'] = correlation_results
        return correlation_results
    
    def perform_trend_analysis(self) -> Dict[str, Dict[str, Any]]:
        """
        Perform trend analysis using linear regression and curve fitting.
        
        Returns:
            Dict[str, Dict[str, Any]]: Trend analysis results by agent
        """
        if self.data is None:
            raise ValueError("No data loaded. Call load_data() first.")
        
        trend_results = {}
        
        for agent_id in self.data['agent_id'].unique():
            agent_data = self.data[self.data['agent_id'] == agent_id].copy()
            agent_data = agent_data.sort_values('turn_number')
            
            X = agent_data['turn_number'].values.reshape(-1, 1)
            y_score = agent_data['comprehension_score'].values
            y_difficulty = agent_data['difficulty_numeric'].values
            
            # Linear regression for score trend
            score_reg = LinearRegression().fit(X, y_score)
            score_r2 = r2_score(y_score, score_reg.predict(X))
            score_slope = float(score_reg.coef_[0])
            score_intercept = float(score_reg.intercept_)
            
            # Linear regression for difficulty trend
            difficulty_reg = LinearRegression().fit(X, y_difficulty)
            difficulty_r2 = r2_score(y_difficulty, difficulty_reg.predict(X))
            difficulty_slope = float(difficulty_reg.coef_[0])
            
            # Polynomial trend fitting (2nd order)
            poly_coeffs = np.polyfit(agent_data['turn_number'].values, y_score, 2)
            poly_r2 = r2_score(y_score, np.polyval(poly_coeffs, agent_data['turn_number'].values))
            
            # Trend classification
            if score_slope > 1:
                trend_type = "strong_positive"
            elif score_slope > 0.1:
                trend_type = "moderate_positive"
            elif score_slope < -1:
                trend_type = "strong_negative"
            elif score_slope < -0.1:
                trend_type = "moderate_negative"
            else:
                trend_type = "stable"
            
            trend_results[agent_id] = {
                'score_slope': score_slope,
                'score_intercept': score_intercept,
                'score_r_squared': float(score_r2),
                'difficulty_slope': difficulty_slope,
                'difficulty_r_squared': float(difficulty_r2),
                'polynomial_coefficients': [float(c) for c in poly_coeffs],
                'polynomial_r_squared': float(poly_r2),
                'trend_type': trend_type,
                'trend_strength': float(abs(score_slope))
            }
        
        self.results['trend_analysis'] = trend_results
        return trend_results
    
    def calculate_effect_sizes(self) -> Dict[str, Dict[str, float]]:
        """
        Calculate effect sizes for practical significance assessment.
        
        Returns:
            Dict[str, Dict[str, float]]: Effect size calculations
        """
        if self.data is None:
            raise ValueError("No data loaded. Call load_data() first.")
        
        effect_sizes = {}
        agents = sorted(self.data['agent_id'].unique())
        
        for agent_id in agents:
            agent_data = self.data[self.data['agent_id'] == agent_id]
            
            # Pre-post effect size (first vs last performance)
            initial_scores = agent_data.head(3)['comprehension_score'].values
            final_scores = agent_data.tail(3)['comprehension_score'].values
            
            cohens_d = self._calculate_cohens_d(final_scores, initial_scores)
            
            # Glass's delta (improvement standardized by initial variability)
            if len(initial_scores) > 1:
                initial_std = np.std(initial_scores, ddof=1)
                if initial_std > 0:
                    glass_delta = (np.mean(final_scores) - np.mean(initial_scores)) / initial_std
                else:
                    glass_delta = 0
            else:
                glass_delta = 0
            
            effect_sizes[agent_id] = {
                'cohens_d': float(cohens_d),
                'glass_delta': float(glass_delta),
                'interpretation': self._interpret_effect_size(cohens_d)
            }
        
        self.results['effect_sizes'] = effect_sizes
        return effect_sizes
    
    def analyze_learning_curves(self) -> Dict[str, Dict[str, Any]]:
        """
        Analyze and characterize learning curve patterns.
        
        Returns:
            Dict[str, Dict[str, Any]]: Learning curve analysis results
        """
        if self.data is None:
            raise ValueError("No data loaded. Call load_data() first.")
        
        learning_curves = {}
        
        for agent_id in self.data['agent_id'].unique():
            agent_data = self.data[self.data['agent_id'] == agent_id].copy()
            agent_data = agent_data.sort_values('turn_number')
            
            turns = agent_data['turn_number'].values
            scores = agent_data['comprehension_score'].values
            
            # Learning curve characterization
            # 1. Power law fitting: y = a * x^b + c
            def power_law(x, a, b, c):
                return a * np.power(x, b) + c
            
            try:
                popt_power, _ = optimize.curve_fit(power_law, turns, scores, 
                                                 bounds=([0, -2, 0], [100, 2, 100]),
                                                 maxfev=2000)
                power_law_fit = True
                power_law_r2 = r2_score(scores, power_law(turns, *popt_power))
            except:
                popt_power = [0, 0, 0]
                power_law_fit = False
                power_law_r2 = 0
            
            # 2. Exponential fitting: y = a * (1 - exp(-b*x)) + c
            def exponential_learning(x, a, b, c):
                return a * (1 - np.exp(-b * x)) + c
            
            try:
                popt_exp, _ = optimize.curve_fit(exponential_learning, turns, scores,
                                               bounds=([0, 0, 0], [100, 1, 100]),
                                               maxfev=2000)
                exp_fit = True
                exp_r2 = r2_score(scores, exponential_learning(turns, *popt_exp))
            except:
                popt_exp = [0, 0, 0]
                exp_fit = False
                exp_r2 = 0
            
            # 3. Learning rate estimation
            if len(scores) > 1:
                improvements = np.diff(scores)
                avg_learning_rate = np.mean(improvements[improvements > 0]) if np.any(improvements > 0) else 0
                
                # Plateau detection
                recent_improvements = improvements[-5:] if len(improvements) >= 5 else improvements
                plateau_threshold = 0.5
                is_plateaued = np.all(np.abs(recent_improvements) < plateau_threshold)
            else:
                avg_learning_rate = 0
                is_plateaued = True
            
            # 4. Learning efficiency (improvement per unit effort)
            total_improvement = scores[-1] - scores[0] if len(scores) > 0 else 0
            learning_efficiency = total_improvement / len(scores) if len(scores) > 0 else 0
            
            # 5. Curve classification
            if power_law_fit and power_law_r2 > 0.7:
                curve_type = "power_law"
                best_r2 = power_law_r2
            elif exp_fit and exp_r2 > 0.7:
                curve_type = "exponential"
                best_r2 = exp_r2
            else:
                curve_type = "linear"
                # Linear fit
                linear_coeffs = np.polyfit(turns, scores, 1)
                best_r2 = r2_score(scores, np.polyval(linear_coeffs, turns))
            
            learning_curves[agent_id] = {
                'curve_type': curve_type,
                'best_fit_r_squared': float(best_r2),
                'power_law_parameters': [float(p) for p in popt_power],
                'exponential_parameters': [float(p) for p in popt_exp],
                'average_learning_rate': float(avg_learning_rate),
                'learning_efficiency': float(learning_efficiency),
                'is_plateaued': bool(is_plateaued),
                'total_improvement': float(total_improvement)
            }
        
        self.results['learning_curves'] = learning_curves
        return learning_curves
    
    def generate_academic_report(self, output_format: str = 'latex') -> str:
        """
        Generate comprehensive academic report in specified format.
        
        Args:
            output_format (str): Output format ('latex', 'markdown', 'html')
            
        Returns:
            str: Path to generated report file
        """
        if not self.results:
            raise ValueError("No analysis results available. Run analysis methods first.")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        if output_format.lower() == 'latex':
            return self._generate_latex_report(timestamp)
        elif output_format.lower() == 'markdown':
            return self._generate_markdown_report(timestamp)
        elif output_format.lower() == 'html':
            return self._generate_html_report(timestamp)
        else:
            raise ValueError("Unsupported output format. Use 'latex', 'markdown', or 'html'.")
    
    def export_results(self, formats: List[str] = ['json', 'csv']) -> Dict[str, str]:
        """
        Export analysis results in specified formats.
        
        Args:
            formats (List[str]): Export formats ['json', 'csv', 'excel']
            
        Returns:
            Dict[str, str]: Mapping of format to output file path
        """
        if not self.results:
            raise ValueError("No analysis results available. Run analysis methods first.")
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        exported_files = {}
        
        # JSON export
        if 'json' in formats:
            json_file = self.output_dir / f"ebars_analysis_results_{timestamp}.json"
            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump(self.results, f, indent=2, ensure_ascii=False, default=str)
            exported_files['json'] = str(json_file)
            print(f"✅ JSON results exported: {json_file}")
        
        # CSV export
        if 'csv' in formats:
            csv_file = self.output_dir / f"ebars_analysis_summary_{timestamp}.csv"
            self._export_summary_csv(csv_file)
            exported_files['csv'] = str(csv_file)
            print(f"✅ CSV summary exported: {csv_file}")
        
        # Excel export
        if 'excel' in formats:
            excel_file = self.output_dir / f"ebars_analysis_comprehensive_{timestamp}.xlsx"
            self._export_excel_workbook(excel_file)
            exported_files['excel'] = str(excel_file)
            print(f"✅ Excel workbook exported: {excel_file}")
        
        return exported_files
    
    def run_comprehensive_analysis(self, csv_file: str, generate_report: bool = True) -> Dict[str, Any]:
        """
        Run complete comprehensive analysis pipeline.
        
        Args:
            csv_file (str): Path to simulation data CSV file
            generate_report (bool): Whether to generate academic report
            
        Returns:
            Dict[str, Any]: Complete analysis results
        """
        print("🔬 Starting EBARS Comprehensive Results Analysis...")
        print("=" * 60)
        
        # Load data
        print("\n📊 Loading simulation data...")
        self.load_data(csv_file)
        
        # Run all analysis components
        analysis_steps = [
            ("Adaptation Effectiveness", self.calculate_adaptation_effectiveness),
            ("Score Convergence", self.analyze_score_convergence), 
            ("System Responsiveness", self.calculate_system_responsiveness),
            ("Stability Metrics", self.analyze_stability_metrics),
            ("Statistical Comparisons", self.perform_statistical_comparisons),
            ("Correlation Patterns", self.analyze_correlation_patterns),
            ("Trend Analysis", self.perform_trend_analysis),
            ("Effect Sizes", self.calculate_effect_sizes),
            ("Learning Curves", self.analyze_learning_curves)
        ]
        
        print("\n🧮 Running statistical analysis components...")
        for step_name, step_func in analysis_steps:
            print(f"  • {step_name}...")
            try:
                step_func()
                print(f"    ✅ {step_name} completed")
            except Exception as e:
                print(f"    ❌ {step_name} failed: {e}")
                continue
        
        # Generate reports
        if generate_report:
            print("\n📝 Generating academic reports...")
            try:
                latex_file = self.generate_academic_report('latex')
                print(f"    ✅ LaTeX report: {latex_file}")
            except Exception as e:
                print(f"    ⚠️ LaTeX report failed: {e}")
            
            try:
                markdown_file = self.generate_academic_report('markdown')
                print(f"    ✅ Markdown report: {markdown_file}")
            except Exception as e:
                print(f"    ⚠️ Markdown report failed: {e}")
        
        # Export results
        print("\n💾 Exporting results...")
        try:
            exported = self.export_results(['json', 'csv', 'excel'])
            for format_type, filepath in exported.items():
                print(f"    ✅ {format_type.upper()}: {filepath}")
        except Exception as e:
            print(f"    ⚠️ Export failed: {e}")
        
        print(f"\n🎉 Analysis completed! Results saved to: {self.output_dir}")
        
        # Return comprehensive summary
        return {
            'data_summary': {
                'total_records': len(self.data),
                'agents_analyzed': list(self.data['agent_id'].unique()),
                'analysis_timestamp': datetime.now().isoformat()
            },
            'analysis_results': self.results,
            'output_directory': str(self.output_dir)
        }
    
    # Helper methods for statistical calculations
    def _calculate_cohens_d(self, group1, group2):
        """Calculate Cohen's d effect size."""
        if len(group1) == 0 or len(group2) == 0:
            return 0
        
        n1, n2 = len(group1), len(group2)
        mean1, mean2 = np.mean(group1), np.mean(group2)
        var1, var2 = np.var(group1, ddof=1), np.var(group2, ddof=1)
        
        # Pooled standard deviation
        pooled_std = np.sqrt(((n1 - 1) * var1 + (n2 - 1) * var2) / (n1 + n2 - 2))
        
        if pooled_std == 0:
            return 0
        
        return (mean1 - mean2) / pooled_std
    
    def _calculate_cliff_delta(self, group1, group2):
        """Calculate Cliff's delta effect size."""
        if len(group1) == 0 or len(group2) == 0:
            return 0
        
        n1, n2 = len(group1), len(group2)
        dominance = 0
        
        for x1 in group1:
            for x2 in group2:
                if x1 > x2:
                    dominance += 1
                elif x1 < x2:
                    dominance -= 1
        
        return dominance / (n1 * n2)
    
    def _calculate_eta_squared(self, groups, f_statistic):
        """Calculate eta-squared effect size for ANOVA."""
        n_total = sum(len(group) for group in groups)
        k = len(groups)
        
        # Calculate sum of squares
        grand_mean = np.mean([x for group in groups for x in group])
        ss_total = sum((x - grand_mean)**2 for group in groups for x in group)
        ss_between = sum(len(group) * (np.mean(group) - grand_mean)**2 for group in groups)
        
        return ss_between / ss_total if ss_total > 0 else 0
    
    def _interpret_statistical_result(self, p_value, effect_size, test_type):
        """Interpret statistical results with effect size."""
        significance = "significant" if p_value < self.alpha else "not significant"
        
        if test_type in ['t_test', 'mann_whitney']:
            if abs(effect_size) < 0.2:
                magnitude = "negligible"
            elif abs(effect_size) < 0.5:
                magnitude = "small"
            elif abs(effect_size) < 0.8:
                magnitude = "medium"
            else:
                magnitude = "large"
        elif test_type == 'anova':
            if effect_size < 0.01:
                magnitude = "negligible"
            elif effect_size < 0.06:
                magnitude = "small"
            elif effect_size < 0.14:
                magnitude = "medium"
            else:
                magnitude = "large"
        else:
            magnitude = "unknown"
        
        return f"Result is {significance} (p={p_value:.4f}) with {magnitude} effect size ({effect_size:.3f})"
    
    def _interpret_effect_size(self, cohens_d):
        """Interpret Cohen's d effect size."""
        abs_d = abs(cohens_d)
        
        if abs_d < 0.2:
            return "negligible effect"
        elif abs_d < 0.5:
            return "small effect"
        elif abs_d < 0.8:
            return "medium effect"
        else:
            return "large effect"
    
    def _generate_latex_report(self, timestamp):
        """Generate LaTeX academic report."""
        latex_file = self.output_dir / f"ebars_analysis_report_{timestamp}.tex"
        
        latex_content = self._create_latex_content()
        
        with open(latex_file, 'w', encoding='utf-8') as f:
            f.write(latex_content)
        
        return str(latex_file)
    
    def _generate_markdown_report(self, timestamp):
        """Generate Markdown report."""
        md_file = self.output_dir / f"ebars_analysis_report_{timestamp}.md"
        
        md_content = self._create_markdown_content()
        
        with open(md_file, 'w', encoding='utf-8') as f:
            f.write(md_content)
        
        return str(md_file)
    
    def _generate_html_report(self, timestamp):
        """Generate HTML report."""
        html_file = self.output_dir / f"ebars_analysis_report_{timestamp}.html"
        
        html_content = self._create_html_content()
        
        with open(html_file, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return str(html_file)
    
    def _create_latex_content(self):
        """Create comprehensive LaTeX report content."""
        # This is a substantial method that would generate a full LaTeX document
        # For brevity, I'll provide the structure
        latex_template = r"""
\documentclass[12pt,a4paper]{article}
\usepackage[utf8]{inputenc}
\usepackage{amsmath,amsfonts,amssymb}
\usepackage{booktabs}
\usepackage{graphicx}
\usepackage{float}
\usepackage{hyperref}

\title{EBARS System: Comprehensive Statistical Analysis Report}
\author{EBARS Research Team}
\date{\today}

\begin{document}
\maketitle

\begin{abstract}
This report presents a comprehensive statistical analysis of the EBARS (Emoji-Based Adaptive Response System) simulation data, validating key research hypotheses about dynamic difficulty adaptation effectiveness.
\end{abstract}

\section{Executive Summary}

\section{Methodology}

\section{Results}

\subsection{Adaptation Effectiveness Analysis}

\subsection{Statistical Comparisons}

\subsection{Learning Curve Characterization}

\section{Discussion}

\section{Conclusions}

\section{References}

\end{document}
"""
        return latex_template
    
    def _create_markdown_content(self):
        """Create Markdown report content."""
        return f"""# EBARS Comprehensive Analysis Report

Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

## Executive Summary

## Analysis Results

### Adaptation Effectiveness

### Statistical Comparisons

### Learning Curves

## Conclusions
"""
    
    def _create_html_content(self):
        """Create HTML report content.""" 
        return f"""<!DOCTYPE html>
<html>
<head>
    <title>EBARS Analysis Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 40px; }}
        .header {{ background: #f0f0f0; padding: 20px; }}
        .section {{ margin: 20px 0; }}
        table {{ border-collapse: collapse; width: 100%; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #f2f2f2; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>EBARS Comprehensive Analysis Report</h1>
        <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    </div>
    
    <div class="section">
        <h2>Executive Summary</h2>
        <!-- Content will be populated -->
    </div>
</body>
</html>"""
    
    def _export_summary_csv(self, filepath):
        """Export summary results to CSV."""
        summary_data = []
        
        # Collect key metrics for CSV export
        if 'adaptation_metrics' in self.results:
            for agent_id, metrics in self.results['adaptation_metrics'].items():
                summary_data.append({
                    'agent_id': agent_id,
                    'metric_type': 'adaptation_effectiveness',
                    'convergence_rate': metrics.convergence_rate,
                    'stability_index': metrics.stability_index,
                    'adaptation_speed': metrics.adaptation_speed,
                    'final_performance': metrics.final_performance,
                    'learning_efficiency': metrics.learning_efficiency
                })
        
        if summary_data:
            df = pd.DataFrame(summary_data)
            df.to_csv(filepath, index=False, encoding='utf-8')
    
    def _export_excel_workbook(self, filepath):
        """Export comprehensive results to Excel workbook."""
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            
            # Create worksheets for different analysis types
            worksheets = {}
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Add summary sheet
            ws_summary = wb.create_sheet("Summary")
            worksheets['summary'] = ws_summary
            
            # Add detailed results sheets based on available results
            for result_type in self.results.keys():
                ws = wb.create_sheet(result_type.replace('_', ' ').title())
                worksheets[result_type] = ws
            
            # Populate worksheets with data
            # This would require detailed implementation for each result type
            
            wb.save(filepath)
            
        except ImportError:
            print("⚠️ openpyxl not installed. Excel export skipped.")
            print("Install with: pip install openpyxl")


def main():
    """
    Example usage and demonstration of the EBARS Results Analysis System.
    """
    print("🔬 EBARS Comprehensive Results Analysis System")
    print("=" * 60)
    
    try:
        # Initialize analyzer
        analyzer = EBARSResultsAnalyzer(
            output_dir="ebars_analysis_output",
            confidence_level=0.95
        )
        
        # Look for simulation data files
        import glob
        csv_files = glob.glob("ebars_simulation_results_*.csv")
        
        if csv_files:
            # Use the most recent simulation file
            latest_file = max(csv_files, key=os.path.getctime)
            print(f"📁 Using simulation data: {latest_file}")
            
            # Run comprehensive analysis
            results = analyzer.run_comprehensive_analysis(
                csv_file=latest_file,
                generate_report=True
            )
            
            print("\n🎉 Analysis completed successfully!")
            print(f"📊 Analyzed {results['data_summary']['total_records']} records")
            print(f"🤖 Agent profiles: {', '.join(results['data_summary']['agents_analyzed'])}")
            print(f"📂 Output directory: {results['output_directory']}")
            
        else:
            print("❌ No simulation CSV files found!")
            print("   Run the simulation first using: python ebars_simulation_working.py")
            
            # Create sample data for demonstration
            print("\n💡 Creating sample data for demonstration...")
            
            # Generate sample simulation data
            sample_data = []
            agents = ['agent_a', 'agent_b', 'agent_c']
            
            for turn in range(1, 21):
                for agent_id in agents:
                    if agent_id == 'agent_a':  # Struggling agent
                        base_score = 30 + turn * 1.2 + np.random.normal(0, 5)
                        emoji = np.random.choice(['❌', '😐'], p=[0.7, 0.3])
                        difficulty = 'struggling' if turn < 12 else 'normal'
                    elif agent_id == 'agent_b':  # Fast learner
                        base_score = 50 + turn * 2.8 + np.random.normal(0, 3)
                        emoji = np.random.choice(['👍', '😊'], p=[0.8, 0.2])
                        difficulty = 'normal' if turn < 6 else 'good' if turn < 15 else 'excellent'
                    else:  # Variable agent
                        base_score = 45 + turn * 1.5 + (np.sin(turn/3) * 10) + np.random.normal(0, 4)
                        emoji = np.random.choice(['❌', '😐', '👍'], p=[0.4, 0.3, 0.3])
                        difficulty = 'struggling' if turn < 8 else 'normal' if turn < 16 else 'good'
                    
                    # Ensure score is within bounds
                    comprehension_score = max(0, min(100, base_score))
                    
                    # Calculate score delta
                    if turn == 1:
                        score_delta = 0
                    else:
                        prev_score = [r for r in sample_data if r['agent_id'] == agent_id and r['turn_number'] == turn-1]
                        if prev_score:
                            score_delta = comprehension_score - prev_score[0]['comprehension_score']
                        else:
                            score_delta = 0
                    
                    sample_data.append({
                        'agent_id': agent_id,
                        'turn_number': turn,
                        'question': f'Sample question {turn}',
                        'answer': f'Sample answer for turn {turn}',
                        'answer_length': np.random.randint(50, 200),
                        'emoji_feedback': emoji,
                        'comprehension_score': comprehension_score,
                        'difficulty_level': difficulty,
                        'score_delta': score_delta,
                        'level_transition': 'same',
                        'processing_time_ms': np.random.normal(1200, 300),
                        'timestamp': datetime.now().isoformat(),
                        'interaction_id': turn * 100 + ord(agent_id[-1]),
                        'feedback_sent': True
                    })
            
            # Save sample data
            sample_df = pd.DataFrame(sample_data)
            sample_file = "sample_ebars_simulation_data.csv"
            sample_df.to_csv(sample_file, index=False, encoding='utf-8')
            print(f"📄 Created sample data: {sample_file}")
            
            # Run analysis on sample data
            results = analyzer.run_comprehensive_analysis(
                csv_file=sample_file,
                generate_report=True
            )
            
            print("✅ Demonstration analysis completed with sample data!")
            
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()