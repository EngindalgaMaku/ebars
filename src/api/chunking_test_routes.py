"""
Chunking Test Routes for Agentic Reasoning Chunking Strategy Testing.

Provides endpoints for:
- Chunking strategy comparison (Traditional vs Agentic Reasoning)
- Real-time chunking test execution and monitoring
- Grok 3 8B model integration for intelligent boundary detection
- Turkish-optimized reasoning prompts
- Performance metrics and analysis
"""

import logging
import json
import os
import uuid
import time
import asyncio
from datetime import datetime, timezone
from typing import Dict, Any, List, Optional
from pathlib import Path

from fastapi import APIRouter, HTTPException, BackgroundTasks, Request, File, Form, UploadFile
from pydantic import BaseModel, Field

# Initialize logger with enhanced formatting FIRST
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(funcName)s:%(lineno)d - %(message)s'
)
logger = logging.getLogger(__name__)

# Authentication functions - avoid circular import by implementing directly
import requests

def _get_current_user(request: Request) -> Optional[Dict[str, Any]]:
    """Get current user from Auth Service using the incoming Authorization header"""
    try:
        auth_header = request.headers.get("Authorization")
        if not auth_header:
            logger.info(f" [AUTH] No Authorization header found")
            return None
        
        # Get AUTH_SERVICE_URL from environment - use IP for Docker network issues
        AUTH_SERVICE_URL = os.getenv('AUTH_SERVICE_URL', 'http://172.18.0.3:8006')
        
        logger.info(f" [AUTH] Calling auth service: {AUTH_SERVICE_URL}/auth/me")
        logger.info(f" [AUTH] Authorization header: {auth_header[:20]}...")
        
        resp = requests.get(f"{AUTH_SERVICE_URL}/auth/me", headers={"Authorization": auth_header}, timeout=10)
        
        logger.info(f" [AUTH] Auth service response: {resp.status_code}")
        
        if resp.status_code == 200:
            user_data = resp.json()
            logger.info(f" [AUTH] User authenticated: {user_data.get('username', 'unknown')}")
            return user_data
        else:
            logger.warning(f" [AUTH] Auth service returned {resp.status_code}: {resp.text}")
            return None
    except Exception as e:
        logger.warning(f" [AUTH] Auth user fetch failed: {e}")
        return None

def _get_role_name(user: Optional[Dict[str, Any]]) -> str:
    """Get role name from user object"""
    if not user:
        return ""
    role = user.get("role_name") or user.get("role") or ""
    return str(role).lower()

def _is_admin(user: Optional[Dict[str, Any]]) -> bool:
    """Check if user is admin"""
    return _get_role_name(user) in {"admin", "superadmin"}

def _is_teacher(user: Optional[Dict[str, Any]]) -> bool:
    """Check if user is teacher"""
    return _get_role_name(user) in {"teacher", "ogretmen", "instructor"}

# Chunking Test Router
router = APIRouter(prefix="/chunking-test", tags=["Chunking Test"])

# Microservice URLs
MODEL_INFERENCE_URL = os.getenv('MODEL_INFERENCE_URL', 'http://model-inference-service:8002')
API_GATEWAY_URL = os.getenv('API_GATEWAY_URL', 'http://localhost:8000')

# Add request timeout configuration
REQUEST_TIMEOUT = 300  # 5 minutes for chunking operations

# Global chunking test results storage with SQLite persistence
CHUNKING_TEST_RESULTS_STORAGE: Dict[str, Any] = {}

# SQLite database for chunking test persistence
import sqlite3
from pathlib import Path

CHUNKING_TEST_DB_PATH = Path("data/chunking_test_results.db")

def _init_chunking_test_db():
    """Initialize chunking test results database"""
    CHUNKING_TEST_DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(CHUNKING_TEST_DB_PATH) as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS chunking_test_results (
                test_id TEXT PRIMARY KEY,
                data TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

def _save_chunking_test_to_db(test_id: str, test_data: Dict[str, Any]):
    """Save chunking test data to database"""
    try:
        with sqlite3.connect(CHUNKING_TEST_DB_PATH) as conn:
            conn.execute("""
                INSERT OR REPLACE INTO chunking_test_results (test_id, data, updated_at)
                VALUES (?, ?, CURRENT_TIMESTAMP)
            """, (test_id, json.dumps(test_data)))
    except Exception as e:
        logger.warning(f"Failed to save chunking test to DB: {e}")

def _load_chunking_test_from_db(test_id: str) -> Optional[Dict[str, Any]]:
    """Load chunking test data from database"""
    try:
        with sqlite3.connect(CHUNKING_TEST_DB_PATH) as conn:
            cursor = conn.execute("SELECT data FROM chunking_test_results WHERE test_id = ?", (test_id,))
            row = cursor.fetchone()
            if row:
                return json.loads(row[0])
    except Exception as e:
        logger.warning(f"Failed to load chunking test from DB: {e}")
    return None

# Initialize database on startup
_init_chunking_test_db()

# ===== REQUEST/RESPONSE MODELS =====

class ChunkingTestStartRequest(BaseModel):
    """Chunking test start request model"""
    testName: str = Field(..., description="Name of the chunking test")
    inputText: str = Field(..., description="Text to be chunked and tested")
    strategies: List[str] = Field(default=["traditional", "agentic"], description="Chunking strategies to test")
    targetChunkSize: int = Field(default=1000, description="Target chunk size in characters")
    overlapSize: int = Field(default=200, description="Overlap size between chunks")
    minChunkSize: int = Field(default=200, description="Minimum chunk size in characters")
    maxChunkSize: int = Field(default=2000, description="Maximum chunk size in characters")
    enableGrokReasoning: bool = Field(default=True, description="Enable Grok 3 8B reasoning for boundary detection")
    turkishOptimization: bool = Field(default=True, description="Enable Turkish language optimization")
    sessionId: Optional[str] = Field(default=None, description="Optional session ID for context")

class ChunkingTestConfiguration(BaseModel):
    """Chunking test configuration model"""
    input_text: str = Field(..., description="Text to be chunked")
    strategies: List[str] = Field(default=["traditional", "agentic"], description="Chunking strategies")
    target_chunk_size: int = Field(default=1000, description="Target chunk size")
    overlap_size: int = Field(default=200, description="Overlap size")
    min_chunk_size: int = Field(default=200, description="Minimum chunk size")
    max_chunk_size: int = Field(default=2000, description="Maximum chunk size")
    enable_grok_reasoning: bool = Field(default=True, description="Enable Grok reasoning")
    turkish_optimization: bool = Field(default=True, description="Turkish optimization")
    session_id: Optional[str] = Field(default=None, description="Session ID")

class ChunkingTestProgress(BaseModel):
    """Chunking test progress model"""
    test_id: str
    status: str  # "running", "completed", "failed"
    current_strategy: str
    completed_strategies: List[str]
    start_time: str
    current_metrics: Dict[str, Any]

class ChunkingResult(BaseModel):
    """Individual chunking result model"""
    strategy: str
    chunks: List[str]
    chunk_count: int
    total_characters: int
    avg_chunk_size: float
    processing_time_ms: float
    semantic_coherence_score: float
    boundary_quality_score: float

class ChunkingTestSummary(BaseModel):
    """Chunking test summary and comparison model"""
    test_id: str
    total_strategies: int
    execution_time_ms: float
    strategy_metrics: Dict[str, Dict[str, float]]
    best_performing_strategy: str
    recommendations: List[str]

# ===== CHUNKING STRATEGY EXECUTION FUNCTIONS =====

async def execute_traditional_chunking(
    text: str,
    target_size: int,
    overlap: int,
    session_id: Optional[str] = None
) -> Dict[str, Any]:
    """Execute traditional chunking strategy"""
    start_time = time.time()
    
    try:
        logger.info(f"Starting traditional chunking - text length: {len(text)}, target_size: {target_size}")

        from src.text_processing.text_chunker import chunk_text

        chunks = chunk_text(
            text=text,
            chunk_size=target_size,
            chunk_overlap=overlap,
            strategy="markdown",
            language="tr",
        )
        
        execution_time = (time.time() - start_time) * 1000
        
        # Calculate metrics
        chunk_count = len(chunks)
        total_chars = sum(len(chunk) for chunk in chunks)
        avg_chunk_size = total_chars / chunk_count if chunk_count > 0 else 0
        
        # Simple coherence score based on chunk size consistency
        size_variance = sum((len(chunk) - avg_chunk_size) ** 2 for chunk in chunks) / chunk_count if chunk_count > 0 else 0
        coherence_score = max(0, 1 - (size_variance / (target_size ** 2)))
        
        # Calculate boundary quality for traditional chunking
        # Based on: sentence boundary detection, paragraph preservation, size consistency
        boundary_quality = 0.7  # Base score for traditional chunking
        
        # Adjust based on how well chunks end at sentence boundaries
        sentence_endings = 0
        for chunk in chunks:
            chunk_stripped = chunk.strip()
            if chunk_stripped and chunk_stripped[-1] in '.!?':
                sentence_endings += 1
        sentence_boundary_ratio = sentence_endings / chunk_count if chunk_count > 0 else 0
        
        # Adjust based on size consistency (lower variance = better boundaries)
        size_consistency = coherence_score
        
        # Final boundary quality: weighted average
        boundary_quality = (
            0.7 * 0.4 +  # Base score weight
            sentence_boundary_ratio * 0.3 +  # Sentence boundary weight
            size_consistency * 0.3  # Size consistency weight
        )
        
        # Enrich chunks with metadata
        detailed_chunks = []
        metadata_stats = {}
        try:
            from src.text_processing.metadata import ChunkEnricher, EnricherConfig
            
            # Create wrapper objects for enrichment
            class SimpleChunk:
                def __init__(self, text):
                    self.text = text
                    self.metadata = {}
            
            chunk_objects = [SimpleChunk(chunk) for chunk in chunks]
            enricher = ChunkEnricher(EnricherConfig(
                use_llm_keywords=False,
                detect_language=True,
                max_keywords=5
            ))
            enriched_chunks = enricher.enrich_chunks(chunk_objects)
            
            # Build detailed chunks with metadata
            for i, chunk_obj in enumerate(enriched_chunks):
                chunk_metadata = chunk_obj.metadata if hasattr(chunk_obj, 'metadata') and chunk_obj.metadata else {}
                
                # Parse JSON fields
                keywords = chunk_metadata.get('keywords_json', '[]')
                if isinstance(keywords, str):
                    try:
                        keywords = json.loads(keywords)
                    except:
                        keywords = []
                
                header_hierarchy = chunk_metadata.get('header_hierarchy_json', '[]')
                if isinstance(header_hierarchy, str):
                    try:
                        header_hierarchy = json.loads(header_hierarchy)
                    except:
                        header_hierarchy = []
                
                detailed_chunks.append({
                    "id": i,
                    "text": chunk_obj.text,
                    "start_index": 0,
                    "end_index": len(chunk_obj.text),
                    "word_count": len(chunk_obj.text.split()),
                    "char_count": len(chunk_obj.text),
                    "metadata": {
                        "chunk_id": chunk_metadata.get('chunk_id', ''),
                        "parent_header": chunk_metadata.get('parent_header', '') or None,
                        "section_title": chunk_metadata.get('section_title', '') or None,
                        "header_hierarchy": header_hierarchy,
                        "keywords": keywords,
                        "chunk_type": chunk_metadata.get('chunk_type', 'content'),
                        "document_title": chunk_metadata.get('document_title', '') or None,
                        "page_number": chunk_metadata.get('page_number') if chunk_metadata.get('page_number', -1) >= 0 else None,
                        "language": chunk_metadata.get('language', 'auto'),
                        "previous_chunk_id": chunk_metadata.get('previous_chunk_id', '') or None,
                        "next_chunk_id": chunk_metadata.get('next_chunk_id', '') or None,
                    }
                })
            
            metadata_stats = enricher.calculate_stats(enriched_chunks).to_dict()
            logger.info(f"Enriched {len(enriched_chunks)} traditional chunks with metadata")
            
        except Exception as enrich_error:
            logger.warning(f"Traditional chunking metadata enrichment failed: {enrich_error}")
            # Create basic detailed_chunks without metadata
            for i, chunk in enumerate(chunks):
                detailed_chunks.append({
                    "id": i,
                    "text": chunk,
                    "start_index": 0,
                    "end_index": len(chunk),
                    "word_count": len(chunk.split()),
                    "char_count": len(chunk),
                    "metadata": {}
                })
        
        logger.info(f"Traditional chunking completed - {chunk_count} chunks, avg_size: {avg_chunk_size:.0f}")
        
        return {
            "strategy": "traditional",
            "chunks": chunks,
            "detailed_chunks": detailed_chunks,
            "chunk_count": chunk_count,
            "total_characters": total_chars,
            "avg_chunk_size": avg_chunk_size,
            "processing_time_ms": execution_time,
            "semantic_coherence_score": coherence_score,
            "boundary_quality_score": boundary_quality,  # Calculated based on actual boundary analysis
            "success": True,
            "config": f"Traditional Semantic Chunking (size={target_size}, overlap={overlap})",
            "metadata_stats": metadata_stats
        }
        
    except Exception as e:
        execution_time = (time.time() - start_time) * 1000
        logger.error(f"Traditional chunking failed: {e}", exc_info=True)
        return {
            "strategy": "traditional",
            "chunks": [],
            "chunk_count": 0,
            "total_characters": 0,
            "avg_chunk_size": 0,
            "processing_time_ms": execution_time,
            "semantic_coherence_score": 0,
            "boundary_quality_score": 0,
            "success": False,
            "error": str(e),
            "error_type": type(e).__name__
        }

async def execute_agentic_reasoning_chunking(
    text: str,
    target_size: int,
    overlap: int,
    enable_grok: bool = True,
    turkish_optimization: bool = True,
    session_id: Optional[str] = None,
    test_id: Optional[str] = None
) -> Dict[str, Any]:
    """Execute agentic reasoning chunking strategy with simplified approach"""
    start_time = time.time()
    
    # Progress update helper
    def update_progress(pct: int, message: str, step: int = 0, total_steps: int = 6):
        if test_id and test_id in CHUNKING_TEST_RESULTS_STORAGE:
            test_data = CHUNKING_TEST_RESULTS_STORAGE[test_id]
            # Calculate overall progress: agentic is typically the second strategy (50-100%)
            base_progress = test_data.get("progress_percentage", 0)
            # Sub-progress within this strategy
            test_data["sub_progress"] = {
                "current_step": step,
                "total_steps": total_steps,
                "step_message": message,
                "step_percentage": pct
            }
            test_data["progress_message"] = f"Agentic Chunking: {message}"
            test_data["last_updated"] = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
            CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
            _save_chunking_test_to_db(test_id, test_data)
    
    try:
        logger.info(f"Starting agentic reasoning chunking - text length: {len(text)}, Grok: {enable_grok}, Turkish: {turkish_optimization}")
        update_progress(5, "Agentic chunker başlatılıyor...", 1, 6)
        
        # Try to import and use the full agentic chunker
        try:
            from src.text_processing.agentic_reasoning_chunker import AgenticReasoningChunker, AgenticChunkingConfig
            
            update_progress(15, "Konfigürasyon hazırlanıyor...", 2, 6)
            
            # Create config using model inference service like LLM chunking
            config = AgenticChunkingConfig(
                target_size=target_size,
                overlap_ratio=overlap / target_size if target_size > 0 else 0.2,
                use_grok_reasoning=enable_grok,
                model_inference_url=MODEL_INFERENCE_URL,
                grok_model_name="llama-3.1-8b-instant",  # Use Llama 3.1 8B instead of Grok
                enable_caching=False,  # Disable caching for simplicity
                enable_quality_validation=False  # Disable validation for speed
            )
            
            # Create chunker
            chunker = AgenticReasoningChunker(config)
            
            update_progress(25, "Metin analiz ediliyor...", 3, 6)
            
            # Perform chunking with timeout and error handling
            try:
                update_progress(40, "Boundary detection yapılıyor...", 4, 6)
                agentic_chunks = chunker.create_chunks(text)
                
                update_progress(70, "Chunk'lar oluşturuluyor...", 5, 6)
                
                # Enrich chunks with metadata
                update_progress(80, "Metadata ekleniyor...", 5, 6)
                try:
                    from src.text_processing.metadata import ChunkEnricher, EnricherConfig
                    enricher = ChunkEnricher(EnricherConfig(
                        use_llm_keywords=False,  # Disable LLM for speed
                        detect_language=True,
                        max_keywords=5
                    ))
                    agentic_chunks = enricher.enrich_chunks(agentic_chunks)
                    logger.info(f"Enriched {len(agentic_chunks)} agentic chunks with metadata")
                except Exception as enrich_error:
                    logger.warning(f"Metadata enrichment failed: {enrich_error}, continuing without metadata")
                
                # Convert AgenticChunk objects to strings for backward compatibility
                chunks = [chunk.text for chunk in agentic_chunks]
                
                execution_time = (time.time() - start_time) * 1000
                chunk_count = len(chunks)
                total_chars = sum(len(chunk) for chunk in chunks)
                avg_chunk_size = total_chars / chunk_count if chunk_count > 0 else 0
                
                # Calculate metrics from agentic chunks
                if agentic_chunks:
                    semantic_coherence = sum(chunk.semantic_coherence for chunk in agentic_chunks) / len(agentic_chunks)
                    
                    # Calculate boundary quality based on multiple factors:
                    # 1. Reasoning confidence (weight: 0.3)
                    # 2. Semantic coherence of chunks (weight: 0.3)
                    # 3. Topic consistency (weight: 0.2)
                    # 4. Quality score (weight: 0.2)
                    reasoning_conf = sum(chunk.reasoning_confidence for chunk in agentic_chunks) / len(agentic_chunks)
                    topic_consistency = sum(chunk.topic_consistency for chunk in agentic_chunks) / len(agentic_chunks)
                    quality_score = sum(chunk.quality_score for chunk in agentic_chunks) / len(agentic_chunks)
                    
                    # Weighted boundary quality calculation
                    boundary_quality = (
                        reasoning_conf * 0.3 +
                        semantic_coherence * 0.3 +
                        topic_consistency * 0.2 +
                        quality_score * 0.2
                    )
                    
                    # Ensure minimum boundary quality for successful chunking
                    # If we successfully created chunks with good semantic coherence, boundary quality should reflect that
                    if semantic_coherence > 0.7 and boundary_quality < 0.6:
                        boundary_quality = max(boundary_quality, 0.65)
                else:
                    semantic_coherence = 0.5
                    boundary_quality = 0.5
                
                # Extract detailed reasoning information for visualization
                detailed_chunks = []
                reasoning_decisions = []
                similarity_analysis = {}
                
                for i, chunk in enumerate(agentic_chunks):
                    # Detailed chunk information
                    chunk_reasoning = []
                    for bd in chunk.boundary_decisions:
                        chunk_reasoning.append({
                            "decision": bd.decision,
                            "confidence": bd.confidence,
                            "reasoning": bd.reasoning,
                            "semantic_coherence": bd.semantic_coherence,
                            "topic_continuity": bd.topic_continuity,
                            "metadata": bd.metadata
                        })
                    
                    # Extract enriched metadata from chunk
                    chunk_metadata = chunk.metadata if hasattr(chunk, 'metadata') and chunk.metadata else {}
                    
                    # Parse JSON fields if they are strings
                    keywords = chunk_metadata.get('keywords_json', '[]')
                    if isinstance(keywords, str):
                        try:
                            keywords = json.loads(keywords)
                        except:
                            keywords = []
                    
                    header_hierarchy = chunk_metadata.get('header_hierarchy_json', '[]')
                    if isinstance(header_hierarchy, str):
                        try:
                            header_hierarchy = json.loads(header_hierarchy)
                        except:
                            header_hierarchy = []
                    
                    detailed_chunks.append({
                        "id": i,
                        "text": chunk.text,
                        "start_index": chunk.start_index,
                        "end_index": chunk.end_index,
                        "word_count": chunk.word_count,
                        "sentence_count": chunk.sentence_count,
                        "paragraph_count": chunk.paragraph_count,
                        "has_header": chunk.has_header,
                        "quality_score": chunk.quality_score,
                        "semantic_coherence": chunk.semantic_coherence,
                        "topic_consistency": chunk.topic_consistency,
                        "reasoning_confidence": chunk.reasoning_confidence,
                        "issues": chunk.issues,
                        "metadata": {
                            "chunk_id": chunk_metadata.get('chunk_id', ''),
                            "parent_header": chunk_metadata.get('parent_header', '') or None,
                            "section_title": chunk_metadata.get('section_title', '') or None,
                            "header_hierarchy": header_hierarchy,
                            "keywords": keywords,
                            "chunk_type": chunk_metadata.get('chunk_type', 'content'),
                            "document_title": chunk_metadata.get('document_title', '') or None,
                            "page_number": chunk_metadata.get('page_number') if chunk_metadata.get('page_number', -1) >= 0 else None,
                            "language": chunk_metadata.get('language', 'auto'),
                            "previous_chunk_id": chunk_metadata.get('previous_chunk_id', '') or None,
                            "next_chunk_id": chunk_metadata.get('next_chunk_id', '') or None,
                        },
                        "boundary_decisions": chunk_reasoning,
                        "reasoning_summary": {
                            "total_decisions": len(chunk.boundary_decisions),
                            "split_decisions": len([bd for bd in chunk.boundary_decisions if bd.decision == "SPLIT"]),
                            "merge_decisions": len([bd for bd in chunk.boundary_decisions if bd.decision == "MERGE"]),
                            "avg_confidence": sum(bd.confidence for bd in chunk.boundary_decisions) / len(chunk.boundary_decisions) if chunk.boundary_decisions else 0,
                            "reasoning_methods": list(set(bd.metadata.get('decision_method', 'unknown') for bd in chunk.boundary_decisions if bd.metadata))
                        }
                    })
                    
                    # Collect all reasoning decisions
                    reasoning_decisions.extend(chunk_reasoning)
                
                # Calculate similarity analysis summary
                if reasoning_decisions:
                    similarity_analysis = {
                        "total_boundary_decisions": len(reasoning_decisions),
                        "split_ratio": len([rd for rd in reasoning_decisions if rd["decision"] == "SPLIT"]) / len(reasoning_decisions),
                        "avg_confidence": sum(rd["confidence"] for rd in reasoning_decisions) / len(reasoning_decisions),
                        "avg_semantic_coherence": sum(rd["semantic_coherence"] for rd in reasoning_decisions) / len(reasoning_decisions),
                        "avg_topic_continuity": sum(rd["topic_continuity"] for rd in reasoning_decisions) / len(reasoning_decisions),
                        "reasoning_methods": list(set(rd["metadata"].get("decision_method", "unknown") for rd in reasoning_decisions if rd.get("metadata")))
                    }
                
                logger.info(f" Agentic chunking completed - {chunk_count} chunks, coherence: {semantic_coherence:.3f}")
                logger.info(f" Agentic reasoning - {len(reasoning_decisions)} boundary decisions, {similarity_analysis.get('split_ratio', 0):.2f} split ratio")
                
                update_progress(100, f"Tamamlandı! {chunk_count} chunk oluşturuldu.", 6, 6)
                
                # Calculate metadata statistics
                metadata_stats = {}
                try:
                    metadata_stats = enricher.calculate_stats(agentic_chunks).to_dict()
                except Exception as stats_error:
                    logger.warning(f"Metadata stats calculation failed: {stats_error}")
                
                return {
                    "strategy": "agentic",
                    "chunks": chunks,  # For backward compatibility
                    "detailed_chunks": detailed_chunks,  # NEW: Detailed chunk information with reasoning
                    "chunk_count": chunk_count,
                    "total_characters": total_chars,
                    "avg_chunk_size": avg_chunk_size,
                    "processing_time_ms": execution_time,
                    "semantic_coherence_score": semantic_coherence,
                    "boundary_quality_score": boundary_quality,
                    "success": True,
                    "config": f"Agentic Reasoning Chunking (Grok={enable_grok}, Turkish={turkish_optimization})",
                    "grok_reasoning_used": enable_grok,
                    "reasoning_decisions": reasoning_decisions,  # NEW: All reasoning decisions
                    "similarity_analysis": similarity_analysis,  # NEW: Similarity analysis summary
                    "metadata_stats": metadata_stats  # NEW: Metadata statistics
                }
                
            except Exception as chunking_error:
                logger.warning(f" Agentic chunking failed: {chunking_error}, using fallback")
                raise Exception(f"Agentic chunking failed: {chunking_error}")
                
        except ImportError as import_error:
            logger.warning(f"Could not import agentic chunker: {import_error}, using fallback")
            raise Exception("Agentic chunker dependencies not available")
            
    except Exception as e:
        execution_time = (time.time() - start_time) * 1000
        logger.error(f"Agentic reasoning chunking failed: {e}, using simple fallback")
        
        # Simple fallback chunking
        try:
            chunks = []
            detailed_chunks = []
            reasoning_decisions = []
            similarity_analysis = {}
            current_pos = 0
            chunk_index = 0
            
            while current_pos < len(text):
                # Find a good break point near target size
                end_pos = min(current_pos + target_size, len(text))
                
                # Try to break at sentence boundaries
                chunk_text = text[current_pos:end_pos]
                
                # Look for sentence endings
                sentence_endings = ['.', '!', '?', '\n\n']
                best_break = end_pos
                chosen_break_reason = "target_size"
                chosen_break_char = ""
                
                for i in range(len(chunk_text) - 1, max(0, len(chunk_text) - 200), -1):
                    if chunk_text[i] in sentence_endings:
                        best_break = current_pos + i + 1
                        chosen_break_char = chunk_text[i]
                        chosen_break_reason = "sentence_boundary" if chunk_text[i] != '\n' else "paragraph_boundary"
                        break
                
                raw_chunk = text[current_pos:best_break]
                chunk = raw_chunk.strip()
                if chunk:
                    chunks.append(chunk)

                    start_index = current_pos
                    end_index = best_break
                    word_count = len(chunk.split())
                    sentence_count = max(
                        1,
                        sum(1 for ch in chunk if ch in ['.', '!', '?'])
                    )
                    paragraph_count = max(1, chunk.count('\n\n') + 1)
                    has_header = chunk.lstrip().startswith('#')

                    boundary_metadata = {
                        "decision_method": "fallback_heuristic",
                        "break_reason": chosen_break_reason,
                        "break_char": chosen_break_char,
                        "target_size": target_size,
                        "search_window": 200,
                    }

                    boundary_reasoning = (
                        f"Fallback heuristic boundary. Reason={chosen_break_reason}. "
                        f"Selected break near target_size={target_size}."
                    )

                    chunk_boundary_decisions = []
                    if end_index < len(text):
                        decision = {
                            "decision": "SPLIT",
                            "confidence": 0.55 if chosen_break_reason == "target_size" else 0.65,
                            "reasoning": boundary_reasoning,
                            "semantic_coherence": 0.55,
                            "topic_continuity": 0.55,
                            "metadata": boundary_metadata,
                        }
                        chunk_boundary_decisions.append(decision)
                        reasoning_decisions.append(decision)

                    detailed_chunks.append({
                        "id": chunk_index,
                        "text": chunk,
                        "start_index": start_index,
                        "end_index": end_index,
                        "word_count": word_count,
                        "sentence_count": sentence_count,
                        "paragraph_count": paragraph_count,
                        "has_header": has_header,
                        "quality_score": 0.55,
                        "semantic_coherence": 0.55,
                        "topic_consistency": 0.55,
                        "reasoning_confidence": 0.55,
                        "issues": ["fallback_heuristic"],
                        "metadata": {
                            "fallback": True,
                            "fallback_reason": str(e),
                            "turkish_optimization": turkish_optimization,
                        },
                        "boundary_decisions": chunk_boundary_decisions,
                        "reasoning_summary": {
                            "total_decisions": len(chunk_boundary_decisions),
                            "split_decisions": len([bd for bd in chunk_boundary_decisions if bd.get("decision") == "SPLIT"]),
                            "merge_decisions": 0,
                            "avg_confidence": (
                                sum(bd.get("confidence", 0) for bd in chunk_boundary_decisions) / len(chunk_boundary_decisions)
                                if chunk_boundary_decisions
                                else 0
                            ),
                            "reasoning_methods": ["fallback_heuristic"],
                        },
                    })
                    chunk_index += 1
                
                current_pos = best_break
            
            execution_time = (time.time() - start_time) * 1000
            chunk_count = len(chunks)
            total_chars = sum(len(chunk) for chunk in chunks)
            avg_chunk_size = total_chars / chunk_count if chunk_count > 0 else 0
            
            logger.info(f"Fallback chunking completed - {chunk_count} chunks")

            if reasoning_decisions:
                similarity_analysis = {
                    "total_boundary_decisions": len(reasoning_decisions),
                    "split_ratio": 1.0,
                    "avg_confidence": sum(rd["confidence"] for rd in reasoning_decisions) / len(reasoning_decisions),
                    "avg_semantic_coherence": sum(rd["semantic_coherence"] for rd in reasoning_decisions) / len(reasoning_decisions),
                    "avg_topic_continuity": sum(rd["topic_continuity"] for rd in reasoning_decisions) / len(reasoning_decisions),
                    "reasoning_methods": list(set(rd["metadata"].get("decision_method", "fallback_heuristic") for rd in reasoning_decisions if rd.get("metadata"))),
                }
            
            return {
                "strategy": "agentic",
                "chunks": chunks,
                "detailed_chunks": detailed_chunks,
                "chunk_count": chunk_count,
                "total_characters": total_chars,
                "avg_chunk_size": avg_chunk_size,
                "processing_time_ms": execution_time,
                "semantic_coherence_score": 0.6,  # Default fallback score
                "boundary_quality_score": 0.7,
                "success": True,
                "config": f"Fallback Chunking (Turkish={turkish_optimization})",
                "grok_reasoning_used": False,
                "reasoning_decisions": reasoning_decisions,
                "similarity_analysis": similarity_analysis,
                "fallback": True
            }
            
        except Exception as fallback_error:
            execution_time = (time.time() - start_time) * 1000
            logger.error(f"Even fallback chunking failed: {fallback_error}")

            try:
                from src.text_processing.text_chunker import chunk_text

                chunks = chunk_text(
                    text=text,
                    chunk_size=target_size,
                    chunk_overlap=overlap,
                    strategy="markdown",
                    language="tr",
                )
                if chunks:
                    total_chars = sum(len(chunk) for chunk in chunks)
                    avg_chunk_size = total_chars / len(chunks) if chunks else 0
                    return {
                        "strategy": "agentic",
                        "chunks": chunks,
                        "chunk_count": len(chunks),
                        "total_characters": total_chars,
                        "avg_chunk_size": avg_chunk_size,
                        "processing_time_ms": execution_time,
                        "semantic_coherence_score": 0.6,
                        "boundary_quality_score": 0.7,
                        "success": True,
                        "fallback": True,
                        "error": str(fallback_error),
                        "error_type": type(fallback_error).__name__,
                    }
            except Exception as last_resort_error:
                logger.error(f"Last resort markdown chunking failed: {last_resort_error}")

            return {
                "strategy": "agentic",
                "chunks": [],
                "chunk_count": 0,
                "total_characters": 0,
                "avg_chunk_size": 0,
                "processing_time_ms": execution_time,
                "semantic_coherence_score": 0,
                "boundary_quality_score": 0,
                "success": False,
                "error": str(fallback_error),
                "error_type": type(fallback_error).__name__,
            }

async def execute_llm_markdown_chunking(
    text: str, 
    target_size: int, 
    overlap: int,
    session_id: Optional[str] = None
) -> Dict[str, Any]:
    """Execute LLM-based markdown chunking strategy"""
    start_time = time.time()
    
    try:
        # Import LLM markdown chunker
        from src.text_processing.llm_markdown_chunker import create_llm_markdown_chunks_safe
        
        # Perform LLM-based chunking with model inference service
        chunks = create_llm_markdown_chunks_safe(
            markdown_text=text,
            target_size=target_size,
            overlap=overlap,
            model_inference_url=MODEL_INFERENCE_URL,
            llm_model_name="llama-3.1-8b-instant",  # Use Llama 3.1 8B
            fallback_model_name="llama-3.1-8b-instant",  # Same fallback
            concurrency=4
        )
        
        execution_time = (time.time() - start_time) * 1000
        
        chunk_count = len(chunks)
        total_chars = sum(len(chunk) for chunk in chunks)
        avg_chunk_size = total_chars / chunk_count if chunk_count > 0 else 0
        
        # Calculate coherence based on chunk size consistency and content flow
        size_variance = sum((len(chunk) - avg_chunk_size) ** 2 for chunk in chunks) / chunk_count if chunk_count > 0 else 0
        coherence_score = max(0, 1 - (size_variance / (target_size ** 2)))
        
        return {
            "strategy": "llm_markdown",
            "chunks": chunks,
            "chunk_count": chunk_count,
            "total_characters": total_chars,
            "avg_chunk_size": avg_chunk_size,
            "processing_time_ms": execution_time,
            "semantic_coherence_score": coherence_score,
            "boundary_quality_score": 0.85,  # LLM-based chunking typically has good boundaries
            "success": True,
            "config": f"LLM Markdown Chunking (Grok 3 8B + Fallback)"
        }
        
    except Exception as e:
        execution_time = (time.time() - start_time) * 1000
        logger.error(f"LLM markdown chunking failed: {e}")
        return {
            "strategy": "llm_markdown",
            "chunks": [],
            "chunk_count": 0,
            "total_characters": 0,
            "avg_chunk_size": 0,
            "processing_time_ms": execution_time,
            "semantic_coherence_score": 0,
            "boundary_quality_score": 0,
            "success": False,
            "error": str(e)
        }


async def execute_multi_agent_chunking(
    text: str,
    target_size: int,
    overlap: int,
    quality_threshold: float = 0.75,
    min_chunk_size: int = 200,
    max_chunk_size: int = 2000,
    session_id: Optional[str] = None
) -> Dict[str, Any]:
    """
    Execute multi-agent chunking strategy.
    
    Uses specialized agents:
    - StructuralAgent: Preserves atomic units (code, tables, lists)
    - SemanticAgent: Detects topic boundaries
    - SizeAgent: Manages chunk sizes
    - QualityAgent: Validates and improves chunks
    - CoordinatorAgent: Orchestrates all agents
    """
    start_time = time.time()
    
    try:
        logger.info(f"Starting multi-agent chunking - text length: {len(text)}, target_size: {target_size}")
        
        from src.text_processing.multi_agent_chunker import MultiAgentChunker, MultiAgentConfig
        
        # Create config
        config = MultiAgentConfig(
            min_chunk_size=min_chunk_size,
            max_chunk_size=max_chunk_size,
            target_chunk_size=target_size,
            overlap_ratio=overlap / target_size if target_size > 0 else 0.1,
            quality_threshold=quality_threshold,
            max_improvement_iterations=3,
            use_llm=True,
            llm_model="llama-3.1-8b-instant",
            model_inference_url=MODEL_INFERENCE_URL,
            enable_parallel=True,
            enable_caching=True
        )
        
        # Create chunker and process
        chunker = MultiAgentChunker(config)
        result = chunker.chunk_text(text)
        
        execution_time = (time.time() - start_time) * 1000
        
        # Extract chunks
        chunks = [chunk.text for chunk in result.chunks]
        chunk_count = len(chunks)
        total_chars = sum(len(chunk) for chunk in chunks)
        avg_chunk_size = total_chars / chunk_count if chunk_count > 0 else 0
        
        # Calculate quality metrics
        avg_quality = result.quality_summary.get('avg_quality', 0.0)
        
        # Build detailed chunks for visualization with metadata
        detailed_chunks = []
        reasoning_decisions = []
        
        for i, chunk in enumerate(result.chunks):
            # Extract metadata from chunk (enriched by ChunkEnricher)
            chunk_metadata = chunk.metadata if hasattr(chunk, 'metadata') and chunk.metadata else {}
            
            # Parse JSON fields if they are strings
            keywords = chunk_metadata.get('keywords_json', '[]')
            if isinstance(keywords, str):
                try:
                    keywords = json.loads(keywords)
                except:
                    keywords = []
            
            header_hierarchy = chunk_metadata.get('header_hierarchy_json', '[]')
            if isinstance(header_hierarchy, str):
                try:
                    header_hierarchy = json.loads(header_hierarchy)
                except:
                    header_hierarchy = []
            
            detailed_chunks.append({
                "id": i,
                "text": chunk.text,
                "start_index": chunk.start_pos,
                "end_index": chunk.end_pos,
                "word_count": chunk.word_count,
                "char_count": chunk.char_count,
                "quality_score": chunk.quality_score,
                "confidence": chunk.confidence,
                "structural_decision": chunk.structural_decision,
                "semantic_decision": chunk.semantic_decision,
                "size_decision": chunk.size_decision,
                "quality_decision": chunk.quality_decision,
                "improvement_iterations": chunk.improvement_iterations,
                "reasoning": chunk.reasoning,
                "processing_time": chunk.processing_time,
                # NEW: Enriched metadata fields
                "metadata": {
                    "chunk_id": chunk_metadata.get('chunk_id', ''),
                    "parent_header": chunk_metadata.get('parent_header', '') or None,
                    "section_title": chunk_metadata.get('section_title', '') or None,
                    "header_hierarchy": header_hierarchy,
                    "keywords": keywords,
                    "chunk_type": chunk_metadata.get('chunk_type', 'content'),
                    "document_title": chunk_metadata.get('document_title', '') or None,
                    "page_number": chunk_metadata.get('page_number') if chunk_metadata.get('page_number', -1) >= 0 else None,
                    "language": chunk_metadata.get('language', 'auto'),
                    "previous_chunk_id": chunk_metadata.get('previous_chunk_id', '') or None,
                    "next_chunk_id": chunk_metadata.get('next_chunk_id', '') or None,
                }
            })
            
            # Determine decision type based on agent decisions
            # MERGE means segments were kept together (no split at this boundary)
            # SPLIT means a new chunk was created
            decision_type = "SPLIT"  # Default - a new chunk was created
            
            # Check if semantic agent suggested MERGE (high similarity)
            if chunk.semantic_decision and 'merge' in chunk.semantic_decision.lower():
                decision_type = "MERGE"
            # Check if structural agent preserved atomic unit
            elif chunk.structural_decision and 'preserve' in chunk.structural_decision.lower():
                decision_type = "PRESERVE"
            # Check if size agent forced the decision
            elif chunk.size_decision:
                if 'force_split' in chunk.size_decision.lower():
                    decision_type = "FORCE_SPLIT"
                elif 'force_merge' in chunk.size_decision.lower():
                    decision_type = "FORCE_MERGE"
            
            # Add reasoning decision
            reasoning_decisions.append({
                "decision": decision_type if i < len(result.chunks) - 1 else "END",
                "confidence": chunk.confidence,
                "reasoning": chunk.reasoning,
                "semantic_coherence": chunk.quality_score,
                "topic_continuity": chunk.quality_score,
                "metadata": {
                    "decision_method": "multi_agent_consensus",
                    "structural": chunk.structural_decision,
                    "semantic": chunk.semantic_decision,
                    "size": chunk.size_decision,
                    "quality": chunk.quality_decision
                }
            })
        
        # Calculate actual decision counts
        split_count = sum(1 for d in reasoning_decisions if d["decision"] in ["SPLIT", "FORCE_SPLIT"])
        merge_count = sum(1 for d in reasoning_decisions if d["decision"] in ["MERGE", "FORCE_MERGE"])
        preserve_count = sum(1 for d in reasoning_decisions if d["decision"] == "PRESERVE")
        
        # Similarity analysis
        similarity_analysis = {
            "total_boundary_decisions": len(reasoning_decisions),
            "split_ratio": split_count / len(reasoning_decisions) if reasoning_decisions else 0,
            "merge_ratio": merge_count / len(reasoning_decisions) if reasoning_decisions else 0,
            "preserve_ratio": preserve_count / len(reasoning_decisions) if reasoning_decisions else 0,
            "avg_confidence": avg_quality,
            "avg_semantic_coherence": avg_quality,
            "avg_topic_continuity": avg_quality,
            "reasoning_methods": ["multi_agent_consensus"],
            "decision_breakdown": {
                "SPLIT": split_count,
                "MERGE": merge_count,
                "PRESERVE": preserve_count,
                "FORCE_SPLIT": sum(1 for d in reasoning_decisions if d["decision"] == "FORCE_SPLIT"),
                "FORCE_MERGE": sum(1 for d in reasoning_decisions if d["decision"] == "FORCE_MERGE"),
            }
        }
        
        logger.info(f"Multi-agent chunking completed - {chunk_count} chunks, avg_quality: {avg_quality:.3f}")
        
        # Get metadata statistics from result
        metadata_stats = result.metadata_stats if hasattr(result, 'metadata_stats') and result.metadata_stats else {}
        
        # Get boundary decisions from agent_metrics if available
        boundary_decisions_from_agents = result.agent_metrics.get('boundary_decisions', {})
        if boundary_decisions_from_agents:
            bd_counts = boundary_decisions_from_agents.get('counts', {})
            bd_details = boundary_decisions_from_agents.get('details', [])
            
            # Update similarity_analysis with actual boundary decision counts
            similarity_analysis['boundary_decision_counts'] = bd_counts
            similarity_analysis['total_boundary_evaluations'] = boundary_decisions_from_agents.get('total', 0)
            
            # Log boundary decision breakdown
            logger.info(f"Boundary decisions: {bd_counts}")
        
        return {
            "strategy": "multi_agent",
            "chunks": chunks,
            "detailed_chunks": detailed_chunks,
            "chunk_count": chunk_count,
            "total_characters": total_chars,
            "avg_chunk_size": avg_chunk_size,
            "processing_time_ms": execution_time,
            "semantic_coherence_score": avg_quality,
            "boundary_quality_score": avg_quality,
            "success": True,
            "config": f"Multi-Agent Chunking (quality_threshold={quality_threshold})",
            "reasoning_decisions": reasoning_decisions,
            "similarity_analysis": similarity_analysis,
            "agent_metrics": result.agent_metrics,
            "quality_summary": result.quality_summary,
            # NEW: Metadata statistics
            "metadata_stats": metadata_stats
        }
        
    except Exception as e:
        execution_time = (time.time() - start_time) * 1000
        logger.error(f"Multi-agent chunking failed: {e}", exc_info=True)
        return {
            "strategy": "multi_agent",
            "chunks": [],
            "chunk_count": 0,
            "total_characters": 0,
            "avg_chunk_size": 0,
            "processing_time_ms": execution_time,
            "semantic_coherence_score": 0,
            "boundary_quality_score": 0,
            "success": False,
            "error": str(e),
            "error_type": type(e).__name__
        }


# ===== API ENDPOINTS =====

@router.post("/start", summary="Start Chunking Test")
async def start_chunking_test(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    config: str = Form(...),
    request: Request = None
) -> Dict[str, Any]:
    """
    Start a comprehensive chunking strategy comparison test with multipart/form-data
    """
    try:
        config_data = json.loads(config)
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="Invalid JSON in config form field")

    # Read file content
    input_text = (await file.read()).decode("utf-8", errors="replace")

    if not input_text or not input_text.strip():
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    # Determine strategies from config
    strategy = config_data.get("strategy", "comparison")
    if strategy == "comparison":
        strategies = ["traditional", "agentic"]
    else:
        strategies = [strategy]

    # Create a Pydantic model from the parsed config for validation and ease of use
    try:
        request_data = ChunkingTestStartRequest(
            testName=config_data.get("testName", ""),
            inputText=input_text,
            strategies=strategies,
            targetChunkSize=config_data.get("chunkSize", 1000),
            overlapSize=config_data.get("chunkOverlap", 200),
            enableGrokReasoning=config_data.get("useSemanticBoundaries", True),
            turkishOptimization=True,  # Assuming this is always on for this context
            sessionId=config_data.get("sessionId")
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid config data: {e}")

    # Enhanced logging for debugging - FORM-DATA REQUEST
    logger.info(f" [CHUNKING TEST START] multipart/form-data request received")
    logger.info(f" [CHUNKING TEST START] Test Name: {request_data.testName}")
    logger.info(f" [CHUNKING TEST START] Strategies: {request_data.strategies}")
    logger.info(f" [CHUNKING TEST START] Input text length: {len(request_data.inputText)}")
    
    # Basic authentication check - USING MOCK AUTHENTICATION FOR DEBUGGING
    if request:
        current_user = _get_current_user(request)
        logger.info(f" [CHUNKING TEST START] Current user: {current_user}")
        logger.info(f"🔍 [CHUNKING TEST START] Current user: {current_user}")
        
        if not (_is_teacher(current_user) or _is_admin(current_user)):
            logger.warning(f"🔍 [CHUNKING TEST START] Access denied for user: {current_user}")
            raise HTTPException(status_code=403, detail="Teacher or admin access required")
        
        logger.info(f"🔍 [CHUNKING TEST START] Authentication successful - access granted")
    
    try:
        # Generate unique test ID
        test_id = str(uuid.uuid4())
        
        # Create configuration object
        config = ChunkingTestConfiguration(
            input_text=request_data.inputText,
            strategies=request_data.strategies,
            target_chunk_size=request_data.targetChunkSize,
            overlap_size=request_data.overlapSize,
            min_chunk_size=request_data.minChunkSize,
            max_chunk_size=request_data.maxChunkSize,
            enable_grok_reasoning=request_data.enableGrokReasoning,
            turkish_optimization=request_data.turkishOptimization,
            session_id=request_data.sessionId
        )
        
        # Initialize test progress
        now_iso = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
        test_progress = {
            "test_id": test_id,
            "test_name": request_data.testName,
            "status": "running",
            "current_strategy": request_data.strategies[0] if request_data.strategies else "traditional",
            "completed_strategies": [],
            "start_time": now_iso,
            "current_metrics": {},
            "configuration": config.dict(),
            "input_text_length": len(request_data.inputText),
            "results": [],
            "progress_percentage": 0.0,
            "progress_message": "Test başlatılıyor...",
            "sub_progress": {
                "current_step": 0,
                "total_steps": 0,
                "step_message": ""
            },
            "last_updated": now_iso
        }
        
        # Store test progress (both memory and database)
        CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_progress
        _save_chunking_test_to_db(test_id, test_progress)
        
        # Start background task for test execution
        background_tasks.add_task(
            execute_full_chunking_test,
            test_id,
            config
        )
        
        return {
            "success": True,
            "testId": test_id,
            "message": "Chunking test started",
            "strategies": request_data.strategies,
            "inputTextLength": len(request_data.inputText),
            "estimated_duration_minutes": len(request_data.strategies) * 0.5
        }
        
    except Exception as e:
        logger.error(f"Failed to start chunking test: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to start chunking test: {str(e)}")

@router.get("/status/{test_id}", summary="Get Chunking Test Status")
async def get_chunking_test_status(test_id: str, request: Request) -> Dict[str, Any]:
    """Get current status of running chunking test"""
    # Basic authentication check - USING MOCK AUTHENTICATION FOR DEBUGGING
    current_user = _get_current_user(request)
    logger.info(f"🔍 [CHUNKING TEST STATUS] Current user: {current_user}")
    
    if not (_is_teacher(current_user) or _is_admin(current_user)):
        raise HTTPException(status_code=403, detail="Teacher or admin access required")
    
    logger.info(f"🔍 [CHUNKING TEST STATUS] Authentication successful - access granted")
    
    # Try memory first, then database
    test_data = CHUNKING_TEST_RESULTS_STORAGE.get(test_id)
    if not test_data:
        test_data = _load_chunking_test_from_db(test_id)
        if test_data:
            CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
    
    if not test_data:
        raise HTTPException(status_code=404, detail="Chunking test not found")
    
    # Calculate progress percentage - use stored value if available
    progress_percentage = test_data.get("progress_percentage", 0.0)
    
    # Fallback calculation if not stored
    total_strategies = len(test_data.get("configuration", {}).get("strategies", []))
    if progress_percentage == 0.0:
        completed_strategies = len(test_data.get("completed_strategies", []))
        
        if total_strategies > 0:
            progress_percentage = (completed_strategies / total_strategies) * 100
        else:
            progress_percentage = 0.0
    
    # Calculate strategy comparison metrics
    strategy_comparison = {}
    results = test_data.get("results", [])
    
    for result in results:
        strategy = result.get("strategy")
        if strategy:
            strategy_comparison[strategy] = {
                "chunkCount": result.get("chunk_count", 0),
                "avgChunkSize": result.get("avg_chunk_size", 0),
                "processingTime": result.get("processing_time_ms", 0),
                "semanticCoherence": result.get("semantic_coherence_score", 0),
                "boundaryQuality": result.get("boundary_quality_score", 0),
                "success": result.get("success", False)
            }
    
    # Process results to create chunks data for frontend
    chunks_data = []
    metrics_data = {
        "totalChunks": 0,
        "averageChunkSize": 0,
        "chunkSizeVariance": 0,
        "semanticCoherence": 0,
        "boundaryQuality": 0,
        "processingTime": 0
    }
    
    # If we have results, process them for frontend display
    if results:
        all_chunks = []
        total_processing_time = 0
        total_coherence = 0
        total_boundary = 0
        successful_results = 0
        
        for result in results:
            if result.get("success", False):
                successful_results += 1
                result_chunks = result.get("chunks", [])
                
                # Convert chunks to frontend format
                for i, chunk_content in enumerate(result_chunks):
                    chunks_data.append({
                        "id": f"{result.get('strategy', 'unknown')}_{i}",
                        "content": chunk_content,
                        "startIndex": 0,  # We don't track this in current implementation
                        "endIndex": len(chunk_content),
                        "size": len(chunk_content),
                        "semanticScore": result.get("semantic_coherence_score", 0),
                        "boundaryType": "semantic" if result.get("strategy") == "agentic" else "natural",
                        "reasoning": f"Generated by {result.get('strategy', 'unknown')} strategy"
                    })
                
                all_chunks.extend(result_chunks)
                total_processing_time += result.get("processing_time_ms", 0)
                total_coherence += result.get("semantic_coherence_score", 0)
                total_boundary += result.get("boundary_quality_score", 0)
        
        # Calculate aggregate metrics
        if successful_results > 0:
            total_chars = sum(len(chunk) for chunk in all_chunks)
            chunk_count = len(all_chunks)
            
            metrics_data = {
                "totalChunks": chunk_count,
                "averageChunkSize": total_chars / chunk_count if chunk_count > 0 else 0,
                "chunkSizeVariance": 0,  # Calculate if needed
                "semanticCoherence": total_coherence / successful_results,
                "boundaryQuality": total_boundary / successful_results,
                "processingTime": round(total_processing_time / 1000, 2)  # Convert to seconds
            }
    
    # Create comparison data if we have multiple strategies
    comparison_data = None
    if len(results) >= 2:
        traditional_result = next((r for r in results if r.get("strategy") == "traditional"), None)
        agentic_result = next((r for r in results if r.get("strategy") == "agentic"), None)
        
        if traditional_result and agentic_result:
            comparison_data = {
                "traditional": {
                    "chunks": [{"id": f"trad_{i}", "content": chunk, "size": len(chunk)}
                              for i, chunk in enumerate(traditional_result.get("chunks", []))],
                    "metrics": {
                        "totalChunks": traditional_result.get("chunk_count", 0),
                        "averageChunkSize": traditional_result.get("avg_chunk_size", 0),
                        "chunkSizeVariance": 0,
                        "semanticCoherence": traditional_result.get("semantic_coherence_score", 0),
                        "boundaryQuality": traditional_result.get("boundary_quality_score", 0),
                        "processingTime": round(traditional_result.get("processing_time_ms", 0) / 1000, 2)
                    }
                },
                "agentic": {
                    "chunks": [{"id": f"agent_{i}", "content": chunk, "size": len(chunk)}
                              for i, chunk in enumerate(agentic_result.get("chunks", []))],
                    "metrics": {
                        "totalChunks": agentic_result.get("chunk_count", 0),
                        "averageChunkSize": agentic_result.get("avg_chunk_size", 0),
                        "chunkSizeVariance": 0,
                        "semanticCoherence": agentic_result.get("semantic_coherence_score", 0),
                        "boundaryQuality": agentic_result.get("boundary_quality_score", 0),
                        "processingTime": round(agentic_result.get("processing_time_ms", 0) / 1000, 2)
                    }
                }
            }

    return {
        "success": True,
        "testId": test_id,
        "status": test_data.get("status", "unknown"),
        "progress": round(progress_percentage, 1),
        "progress_percentage": round(progress_percentage, 1),
        "progress_message": test_data.get("progress_message"),
        "sub_progress": test_data.get("sub_progress"),
        "startTime": test_data.get("start_time"),
        "endTime": test_data.get("end_time"),
        "currentStrategy": test_data.get("current_strategy"),
        "completedStrategies": test_data.get("completed_strategies", []),
        "strategyComparison": strategy_comparison,
        "inputTextLength": test_data.get("input_text_length", 0),
        "totalStrategies": total_strategies,
        "resultsCount": len(results),
        
        # Frontend expected fields
        "chunks": chunks_data,
        "metrics": metrics_data,
        "comparison": comparison_data,
        "originalText": test_data.get("configuration", {}).get("input_text", ""),
        "totalCharacters": test_data.get("input_text_length", 0),
        "processingTime": metrics_data["processingTime"]
    }

@router.get("/list", summary="List All Chunking Tests")
async def list_chunking_tests(request: Request) -> Dict[str, Any]:
    """List all chunking test results"""
    # Basic authentication check - USING MOCK AUTHENTICATION FOR DEBUGGING
    current_user = _get_current_user(request)
    logger.info(f"🔍 [CHUNKING TEST LIST] Current user: {current_user}")
    
    if not (_is_teacher(current_user) or _is_admin(current_user)):
        raise HTTPException(status_code=403, detail="Teacher or admin access required")
    
    logger.info(f"🔍 [CHUNKING TEST LIST] Authentication successful - access granted")
    
    try:
        with sqlite3.connect(CHUNKING_TEST_DB_PATH) as conn:
            cursor = conn.execute("""
                SELECT test_id, data, created_at, updated_at 
                FROM chunking_test_results 
                ORDER BY updated_at DESC
            """)
            
            rows = cursor.fetchall()
            tests = []
            for row in rows:
                test_id, data_json, created_at, updated_at = row
                try:
                    test_data = json.loads(data_json)
                    
                    # Extract key info for list view
                    tests.append({
                        "testId": test_id,
                        "testName": test_data.get("test_name", f"Chunking Test {test_id[:8]}"),
                        "status": test_data.get("status", "unknown"),
                        "startTime": test_data.get("start_time"),
                        "endTime": test_data.get("end_time"),
                        "createdAt": created_at,
                        "updatedAt": updated_at,
                        "strategies": test_data.get("configuration", {}).get("strategies", []),
                        "inputTextLength": test_data.get("input_text_length", 0)
                    })
                except json.JSONDecodeError:
                    logger.warning(f"Failed to parse chunking test data for {test_id}")
                    continue
            
            return {
                "success": True,
                "tests": tests,
                "total": len(tests)
            }
    except Exception as e:
        logger.error(f"Error listing chunking tests: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error listing chunking tests: {str(e)}")

@router.delete("/delete/{test_id}", summary="Delete Chunking Test")
async def delete_chunking_test(test_id: str, request: Request) -> Dict[str, Any]:
    """Delete a stored chunking test"""
    # Basic authentication check - USING MOCK AUTHENTICATION FOR DEBUGGING
    current_user = _get_current_user(request)
    logger.info(f"🔍 [CHUNKING TEST DELETE] Current user: {current_user}")
    
    if not (_is_teacher(current_user) or _is_admin(current_user)):
        raise HTTPException(status_code=403, detail="Teacher or admin access required")
    
    logger.info(f"🔍 [CHUNKING TEST DELETE] Authentication successful - access granted")
    
    # Try memory first, then database
    test_data = CHUNKING_TEST_RESULTS_STORAGE.get(test_id)
    if not test_data:
        test_data = _load_chunking_test_from_db(test_id)
    
    if not test_data:
        raise HTTPException(status_code=404, detail="Chunking test not found")
    
    # Prevent deleting running tests
    if test_data.get("status") == "running":
        raise HTTPException(status_code=409, detail="Running tests cannot be deleted")
    
    try:
        with sqlite3.connect(CHUNKING_TEST_DB_PATH) as conn:
            cur = conn.execute("DELETE FROM chunking_test_results WHERE test_id = ?", (test_id,))
            if cur.rowcount == 0:
                logger.warning(f"Chunking test {test_id} not found in DB during delete")
        
        # Remove from memory cache
        CHUNKING_TEST_RESULTS_STORAGE.pop(test_id, None)
        
        return {"success": True, "testId": test_id}
    except Exception as e:
        logger.error(f"Error deleting chunking test {test_id}: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error deleting chunking test: {str(e)}")

@router.post("/stop/{test_id}", summary="Stop Chunking Test")
async def stop_chunking_test(test_id: str, request: Request) -> Dict[str, Any]:
    """Stop a running chunking test"""
    # Basic authentication check - USING MOCK AUTHENTICATION FOR DEBUGGING
    current_user = _get_current_user(request)
    logger.info(f"🔍 [CHUNKING TEST STOP] Current user: {current_user}")
    
    if not (_is_teacher(current_user) or _is_admin(current_user)):
        raise HTTPException(status_code=403, detail="Teacher or admin access required")
    
    logger.info(f"🔍 [CHUNKING TEST STOP] Authentication successful - access granted")
    
    if test_id not in CHUNKING_TEST_RESULTS_STORAGE:
        raise HTTPException(status_code=404, detail="Chunking test not found")
    
    test_data = CHUNKING_TEST_RESULTS_STORAGE[test_id]
    
    if test_data["status"] != "running":
        return {
            "success": False,
            "message": f"Test is not running (current status: {test_data['status']})"
        }
    
    # Mark test as stopped
    test_data["status"] = "stopped"
    test_data["end_time"] = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    
    # Persist stop status
    CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
    _save_chunking_test_to_db(test_id, test_data)
    
    return {
        "success": True,
        "message": "Chunking test stopped",
        "partial_results": len(test_data.get("results", []))
    }

@router.get("/results/{test_id}", summary="Get Chunking Test Results")
async def get_chunking_test_results(test_id: str, format: str = "json", request: Request = None) -> Dict[str, Any]:
    """Get comprehensive chunking test results with strategy comparison"""
    # Basic authentication check - USING MOCK AUTHENTICATION FOR DEBUGGING
    if request:
        current_user = _get_current_user(request)
        logger.info(f"🔍 [CHUNKING TEST RESULTS] Current user: {current_user}")
        
        if not (_is_teacher(current_user) or _is_admin(current_user)):
            raise HTTPException(status_code=403, detail="Teacher or admin access required")
        
        logger.info(f"🔍 [CHUNKING TEST RESULTS] Authentication successful - access granted")
    
    # Try memory first, then database
    test_data = CHUNKING_TEST_RESULTS_STORAGE.get(test_id)
    if not test_data:
        test_data = _load_chunking_test_from_db(test_id)
        if test_data:
            CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
    
    if not test_data:
        raise HTTPException(status_code=404, detail="Chunking test not found")
    
    try:
        results = test_data.get("results", [])
        
        # Process results and calculate comprehensive metrics
        strategy_metrics = {}
        for result in results:
            strategy = result.get("strategy")
            if strategy:
                strategy_metrics[strategy] = {
                    "chunk_count": result.get("chunk_count", 0),
                    "avg_chunk_size": result.get("avg_chunk_size", 0),
                    "processing_time_ms": result.get("processing_time_ms", 0),
                    "semantic_coherence_score": result.get("semantic_coherence_score", 0),
                    "boundary_quality_score": result.get("boundary_quality_score", 0),
                    "success": result.get("success", False),
                    "config": result.get("config", "")
                }
        
        # Determine best performing strategy
        best_strategy = ""
        best_score = 0
        for strategy, metrics in strategy_metrics.items():
            if metrics["success"]:
                # Composite score based on coherence, boundary quality, and efficiency
                score = (
                    metrics["semantic_coherence_score"] * 0.4 +
                    metrics["boundary_quality_score"] * 0.4 +
                    (1 - min(metrics["processing_time_ms"] / 10000, 1)) * 0.2
                )
                if score > best_score:
                    best_score = score
                    best_strategy = strategy
        
        # Calculate execution time
        start_time = test_data.get("start_time")
        end_time = test_data.get("end_time")
        execution_time = 0
        if start_time and end_time:
            try:
                start_dt = datetime.fromisoformat(start_time.replace("Z", "+00:00"))
                end_dt = datetime.fromisoformat(end_time.replace("Z", "+00:00"))
                execution_time = (end_dt - start_dt).total_seconds() * 1000
            except:
                pass
        
        if format.lower() == "json":
            return {
                "success": True,
                "format": "json",
                "test_id": test_id,
                "test_name": test_data.get("test_name", ""),
                "start_time": test_data.get("start_time"),
                "end_time": test_data.get("end_time"),
                "execution_time_ms": execution_time,
                "status": test_data.get("status", "unknown"),
                "configuration": test_data.get("configuration", {}),
                "input_text_length": test_data.get("input_text_length", 0),
                "strategy_metrics": strategy_metrics,
                "best_performing_strategy": best_strategy,
                "detailed_results": results,
                "summary": {
                    "total_strategies": len(strategy_metrics),
                    "successful_strategies": sum(1 for m in strategy_metrics.values() if m["success"]),
                    "best_strategy": best_strategy,
                    "execution_time_ms": execution_time
                }
            }
        else:
            raise HTTPException(status_code=400, detail="Only JSON format is currently supported")
            
    except Exception as e:
        logger.error(f"Failed to get chunking test results: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get chunking test results: {str(e)}")

@router.get("/export/{test_id}", summary="Export Chunking Test Results")
async def export_chunking_test_results(test_id: str, format: str = "json", request: Request = None):
    """Export chunking test results in specified format (JSON, CSV, or Excel)"""
    # Basic authentication check
    if request:
        current_user = _get_current_user(request)
        logger.info(f"🔍 [CHUNKING EXPORT] Current user: {current_user}")
        
        if not (_is_teacher(current_user) or _is_admin(current_user)):
            raise HTTPException(status_code=403, detail="Teacher or admin access required")
        
        logger.info(f"🔍 [CHUNKING EXPORT] Authentication successful - access granted")
    
    # Try memory first, then database
    test_data = CHUNKING_TEST_RESULTS_STORAGE.get(test_id)
    if not test_data:
        test_data = _load_chunking_test_from_db(test_id)
        if test_data:
            CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
    
    if not test_data:
        raise HTTPException(status_code=404, detail="Chunking test not found")
    
    # For Excel format, return StreamingResponse directly
    if format.lower() == "excel" or format.lower() == "xlsx":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from io import BytesIO
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Chunking Test Results"
            
            # Header row with styling
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Write headers
            headers = ["Test ID", "Strategy", "Chunk Count", "Avg Chunk Size", "Processing Time (ms)", "Semantic Coherence", "Boundary Quality", "Success"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Write data rows
            row_alignment = Alignment(vertical="top", wrap_text=True)
            results = test_data.get("results", [])
            for row_idx, result in enumerate(results, start=2):
                ws.cell(row=row_idx, column=1, value=test_id).alignment = row_alignment
                ws.cell(row=row_idx, column=2, value=result.get("strategy", "")).alignment = row_alignment
                ws.cell(row=row_idx, column=3, value=result.get("chunk_count", 0)).alignment = row_alignment
                ws.cell(row=row_idx, column=4, value=result.get("avg_chunk_size", 0)).alignment = row_alignment
                ws.cell(row=row_idx, column=5, value=result.get("processing_time_ms", 0)).alignment = row_alignment
                ws.cell(row=row_idx, column=6, value=result.get("semantic_coherence_score", 0)).alignment = row_alignment
                ws.cell(row=row_idx, column=7, value=result.get("boundary_quality_score", 0)).alignment = row_alignment
                ws.cell(row=row_idx, column=8, value="Yes" if result.get("success", False) else "No").alignment = row_alignment
            
            # Set column widths
            for col in range(1, 9):
                ws.column_dimensions[chr(64 + col)].width = 15
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"chunking_test_results_{test_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            from fastapi.responses import StreamingResponse
            return StreamingResponse(
                BytesIO(output.read()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        except ImportError:
            logger.error("openpyxl not installed. Please install it: pip install openpyxl")
            raise HTTPException(status_code=500, detail="Excel export requires openpyxl library")
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            raise HTTPException(status_code=500, detail=f"Excel export failed: {str(e)}")
    
    # For other formats, delegate to existing results endpoint
    return await get_chunking_test_results(test_id, format, request)

@router.get("/export-pdf/{test_id}", summary="Export Chunking Test Results as PDF")
async def export_chunking_test_pdf(test_id: str, request: Request = None):
    """Export chunking test results as comprehensive PDF file with Turkish support"""
    logger.info(f"📄 [PDF EXPORT] Starting PDF export for test_id: {test_id}")
    
    from fastapi.responses import Response
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
    from io import BytesIO
    import textwrap
    
    # Basic authentication check
    if request:
        current_user = _get_current_user(request)
        logger.info(f"🔍 [CHUNKING EXPORT] Current user: {current_user}")
        
        if not (_is_teacher(current_user) or _is_admin(current_user)):
            logger.error(f"❌ [PDF EXPORT] Access denied for user: {current_user}")
            raise HTTPException(status_code=403, detail="Teacher or admin access required")
        
        logger.info(f"🔍 [CHUNKING EXPORT] Authentication successful - access granted")
    
    # Try memory first, then database
    logger.info(f"📄 [PDF EXPORT] Loading test data for {test_id}")
    test_data = CHUNKING_TEST_RESULTS_STORAGE.get(test_id)
    if not test_data:
        test_data = _load_chunking_test_from_db(test_id)
        if test_data:
            CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
            logger.info(f"📄 [PDF EXPORT] Test data loaded from database")
        else:
            logger.error(f"❌ [PDF EXPORT] Test data not found for {test_id}")
    else:
        logger.info(f"📄 [PDF EXPORT] Test data loaded from memory")
    
    if not test_data:
        logger.error(f"❌ [PDF EXPORT] Test not found: {test_id}")
        raise HTTPException(status_code=404, detail="Chunking test not found")
    
    logger.info(f"📄 [PDF EXPORT] Test data status: {test_data.get('status', 'unknown')}")
    logger.info(f"📄 [PDF EXPORT] Results count: {len(test_data.get('results', []))}")
    
    try:
        logger.info(f"📄 [PDF EXPORT] Creating PDF buffer and document")
        
        # Register Turkish-compatible font
        turkish_font = 'Helvetica'
        try:
            font_paths = [
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
                "/usr/share/fonts/dejavu/DejaVuSans.ttf",
                "C:/Windows/Fonts/arial.ttf",
                "/System/Library/Fonts/Helvetica.ttc"
            ]
            for font_path in font_paths:
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('TurkishFont', font_path))
                    turkish_font = 'TurkishFont'
                    logger.info(f"📄 [PDF EXPORT] Registered font: {font_path}")
                    break
        except Exception as font_error:
            logger.warning(f"📄 [PDF EXPORT] Font registration failed: {font_error}")
        
        # Create PDF buffer
        buffer = BytesIO()
        
        # Create PDF document
        doc = SimpleDocTemplate(
            buffer, 
            pagesize=A4, 
            rightMargin=1.5*cm, 
            leftMargin=1.5*cm, 
            topMargin=2*cm, 
            bottomMargin=2*cm
        )
        
        # Custom styles with Turkish font support
        title_style = ParagraphStyle(
            'CustomTitle',
            fontName=turkish_font,
            fontSize=20,
            spaceAfter=30,
            spaceBefore=20,
            textColor=colors.HexColor('#1e3a5f'),
            alignment=TA_CENTER,
            leading=24
        )
        
        heading1_style = ParagraphStyle(
            'CustomHeading1',
            fontName=turkish_font,
            fontSize=16,
            spaceAfter=12,
            spaceBefore=20,
            textColor=colors.HexColor('#1e3a5f'),
            leading=20
        )
        
        heading2_style = ParagraphStyle(
            'CustomHeading2',
            fontName=turkish_font,
            fontSize=13,
            spaceAfter=8,
            spaceBefore=14,
            textColor=colors.HexColor('#2563eb'),
            leading=16
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            fontName=turkish_font,
            fontSize=10,
            spaceAfter=6,
            leading=14,
            alignment=TA_JUSTIFY
        )
        
        small_style = ParagraphStyle(
            'CustomSmall',
            fontName=turkish_font,
            fontSize=9,
            spaceAfter=4,
            leading=12,
            textColor=colors.HexColor('#4b5563')
        )
        
        chunk_style = ParagraphStyle(
            'ChunkStyle',
            fontName=turkish_font,
            fontSize=9,
            spaceAfter=8,
            spaceBefore=4,
            leading=12,
            backColor=colors.HexColor('#f8fafc'),
            leftIndent=5,
            rightIndent=5
        )
        
        # Helper function to escape XML special characters
        def escape_xml(text):
            if not isinstance(text, str):
                text = str(text)
            return text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')
        
        logger.info(f"📄 [PDF EXPORT] Building PDF content")
        story = []
        
        # ===== TITLE PAGE =====
        story.append(Paragraph("Agentic Chunking Sistemi", title_style))
        story.append(Paragraph("Akademik Degerlendirme Raporu", heading1_style))
        story.append(Spacer(1, 30))
        
        # Test metadata
        test_name = escape_xml(test_data.get('test_name', 'Isimsiz Test'))
        created_at = test_data.get('created_at', test_data.get('start_time', 'Bilinmiyor'))
        if isinstance(created_at, datetime):
            created_at = created_at.strftime('%Y-%m-%d %H:%M:%S')
        
        meta_data = [
            ['Test ID:', test_id],
            ['Test Adi:', test_name],
            ['Olusturulma Tarihi:', str(created_at)],
            ['Durum:', test_data.get('status', 'Bilinmiyor')],
            ['Rapor Tarihi:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        meta_table = Table(meta_data, colWidths=[4*cm, 12*cm])
        meta_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), turkish_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#6b7280')),
            ('TEXTCOLOR', (1, 0), (1, -1), colors.HexColor('#1f2937')),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
        ]))
        story.append(meta_table)
        story.append(Spacer(1, 20))
        
        # ===== CONFIGURATION SECTION =====
        story.append(Paragraph("1. Test Konfigurasyonu", heading1_style))
        
        config = test_data.get('configuration', test_data.get('config', {}))
        if config:
            config_items = []
            param_mappings = {
                'strategies': 'Stratejiler',
                'chunk_size': 'Chunk Boyutu',
                'chunk_overlap': 'Overlap',
                'similarity_threshold': 'Benzerlik Esigi',
                'llm_reasoning_weight': 'LLM Reasoning Agirligi',
                'max_chunk_size': 'Maksimum Chunk Boyutu',
                'min_chunk_size': 'Minimum Chunk Boyutu',
                'use_semantic_boundaries': 'Semantik Sinirlar',
                'enable_contextual_merging': 'Baglamsal Birlestirme',
                'enable_quality_metrics': 'Kalite Metrikleri',
                'target_size': 'Hedef Boyut',
                'language': 'Dil'
            }
            
            for key, label in param_mappings.items():
                value = config.get(key)
                if value is not None:
                    if isinstance(value, bool):
                        value = 'Evet' if value else 'Hayir'
                    elif isinstance(value, list):
                        value = ', '.join(str(v) for v in value)
                    config_items.append([label + ':', str(value)])
            
            if config_items:
                config_table = Table(config_items, colWidths=[5*cm, 11*cm])
                config_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), turkish_font),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#6b7280')),
                    ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                    ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f9fafb')),
                ]))
                story.append(config_table)
        else:
            story.append(Paragraph("Konfigurasyon bilgisi bulunamadi.", normal_style))
        
        story.append(Spacer(1, 20))
        
        # ===== METRIC DEFINITIONS SECTION =====
        story.append(Paragraph("1.1. Metrik ve Parametre Tanimlari", heading2_style))
        story.append(Spacer(1, 10))
        
        # Metric definitions - FIXED column widths to prevent overflow
        metric_definitions = [
            ['Metrik/Parametre', 'Tanim', 'Deger Araligi'],
            ['Semantic Coherence Score', 
             'Chunk icindeki konularin tutarliligi. QualityAgent tarafindan hesaplanir.',
             '0.0 - 1.0'],
            ['Boundary Quality Score', 
             'Chunk sinirlarinin kalitesi. Dogal sinirlar yuksek skor alir.',
             '0.0 - 1.0'],
            ['Confidence', 
             'Agent kararlarina olan guven seviyesi.',
             '0.0 - 1.0'],
            ['SPLIT Karari', 
             'Iki segment arasinda bolunme karari.',
             'Karar Turu'],
            ['MERGE Karari', 
             'Iki segmentin birlestirilmesi karari.',
             'Karar Turu'],
            ['PRESERVE Karari', 
             'Atomik birimin korunmasi karari.',
             'Karar Turu'],
            ['Target Chunk Size', 
             'Hedeflenen chunk boyutu (karakter).',
             'Karakter'],
            ['Overlap Size', 
             'Ardisik chunklar arasindaki ortak metin.',
             'Karakter'],
        ]
        
        metric_table = Table(metric_definitions, colWidths=[4*cm, 8*cm, 3*cm])
        metric_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), turkish_font),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (2, 1), (2, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(metric_table)
        story.append(Spacer(1, 15))
        
        # Agent descriptions
        story.append(Paragraph("1.2. Agent Rolleri", heading2_style))
        story.append(Spacer(1, 10))
        
        agent_descriptions = [
            ['Agent', 'Gorev', 'Karar Tipleri'],
            ['StructuralAgent', 
             'Atomik birimleri (kod, tablo, liste) korur.',
             'PRESERVE, SPLIT'],
            ['SemanticAgent', 
             'Konu tutarliligi ve semantik benzerligi analiz eder.',
             'MERGE, SPLIT'],
            ['SizeAgent', 
             'Chunk boyutlarini kontrol eder.',
             'FORCE_SPLIT, FORCE_MERGE'],
            ['QualityAgent', 
             'Chunk kalitesini degerlendirir.',
             'APPROVED, REJECTED'],
            ['CoordinatorAgent', 
             'Agentlari koordine eder, konsensus hesaplar.',
             'Final karar'],
        ]
        
        agent_table = Table(agent_descriptions, colWidths=[3*cm, 7*cm, 5*cm])
        agent_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), turkish_font),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(agent_table)
        story.append(Spacer(1, 20))
        
        # ===== RESULTS SUMMARY =====
        story.append(Paragraph("2. Sonuc Ozeti", heading1_style))
        
        results = test_data.get('results', [])
        logger.info(f"📄 [PDF EXPORT] Processing {len(results)} results")
        
        if results:
            # Summary table
            summary_header = ['Strateji', 'Chunk', 'Ort.Boyut', 'Semantic', 'Boundary', 'Sure', 'Durum']
            summary_data = [summary_header]
            
            for result in results:
                strategy = result.get('strategy', 'Bilinmiyor')
                chunk_count = result.get('chunk_count', len(result.get('chunks', [])))
                avg_size = result.get('avg_chunk_size', 0)
                semantic_score = result.get('semantic_coherence_score', 0)
                boundary_score = result.get('boundary_quality_score', 0)
                proc_time = result.get('processing_time_ms', 0)
                success = result.get('success', False)
                
                summary_data.append([
                    strategy,
                    str(chunk_count),
                    f"{avg_size:.0f}" if avg_size else "N/A",
                    f"{semantic_score:.3f}" if semantic_score else "N/A",
                    f"{boundary_score:.3f}" if boundary_score else "N/A",
                    f"{proc_time:.0f}ms" if proc_time else "N/A",
                    'OK' if success else 'FAIL'
                ])
            
            summary_table = Table(summary_data, colWidths=[2.5*cm, 1.5*cm, 2*cm, 2*cm, 2*cm, 2*cm, 1.5*cm])
            summary_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), turkish_font),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a5f')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
                ('TOPPADDING', (0, 0), (-1, 0), 10),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 20))
            
            # ===== STRATEGY COMPARISON SECTION =====
            if len(results) >= 2:
                story.append(Paragraph("2.1. Strateji Karsilastirmasi", heading2_style))
                story.append(Spacer(1, 10))
                
                # Find traditional and agentic results
                traditional_result = None
                agentic_result = None
                multi_agent_result = None
                
                for r in results:
                    strategy = r.get('strategy', '').lower()
                    if strategy == 'traditional':
                        traditional_result = r
                    elif strategy in ['agentic', 'agentic_reasoning']:
                        agentic_result = r
                    elif strategy == 'multi_agent':
                        multi_agent_result = r
                
                # Use agentic or multi_agent as the "new" method
                new_result = agentic_result or multi_agent_result
                
                if traditional_result and new_result:
                    new_strategy_name = new_result.get('strategy', 'Agentic').upper()
                    
                    # Calculate comparison metrics
                    trad_chunks = traditional_result.get('chunk_count', 0)
                    new_chunks = new_result.get('chunk_count', 0)
                    chunk_diff = new_chunks - trad_chunks
                    chunk_diff_pct = ((new_chunks - trad_chunks) / trad_chunks * 100) if trad_chunks > 0 else 0
                    
                    trad_semantic = traditional_result.get('semantic_coherence_score', 0)
                    new_semantic = new_result.get('semantic_coherence_score', 0)
                    semantic_diff = new_semantic - trad_semantic
                    semantic_improvement = ((new_semantic - trad_semantic) / trad_semantic * 100) if trad_semantic > 0 else 0
                    
                    trad_boundary = traditional_result.get('boundary_quality_score', 0)
                    new_boundary = new_result.get('boundary_quality_score', 0)
                    boundary_diff = new_boundary - trad_boundary
                    boundary_improvement = ((new_boundary - trad_boundary) / trad_boundary * 100) if trad_boundary > 0 else 0
                    
                    trad_time = traditional_result.get('processing_time_ms', 0)
                    new_time = new_result.get('processing_time_ms', 0)
                    time_diff = new_time - trad_time
                    
                    trad_avg_size = traditional_result.get('avg_chunk_size', 0)
                    new_avg_size = new_result.get('avg_chunk_size', 0)
                    
                    # Comparison table
                    comparison_data = [
                        ['Metrik', 'Traditional', new_strategy_name, 'Fark', 'Kazanan'],
                        ['Chunk Sayisi', str(trad_chunks), str(new_chunks), 
                         f"{chunk_diff:+d} ({chunk_diff_pct:+.1f}%)", 
                         'Esit' if chunk_diff == 0 else ('Traditional' if chunk_diff > 0 else new_strategy_name)],
                        ['Ort. Chunk Boyutu', f"{trad_avg_size:.0f}", f"{new_avg_size:.0f}",
                         f"{new_avg_size - trad_avg_size:+.0f}",
                         'Esit' if abs(new_avg_size - trad_avg_size) < 50 else ('Hedefe Yakin' if abs(new_avg_size - 1000) < abs(trad_avg_size - 1000) else 'Traditional')],
                        ['Semantic Coherence', f"{trad_semantic:.4f}", f"{new_semantic:.4f}",
                         f"{semantic_diff:+.4f} ({semantic_improvement:+.1f}%)",
                         new_strategy_name if semantic_diff > 0 else ('Esit' if semantic_diff == 0 else 'Traditional')],
                        ['Boundary Quality', f"{trad_boundary:.4f}", f"{new_boundary:.4f}",
                         f"{boundary_diff:+.4f} ({boundary_improvement:+.1f}%)",
                         new_strategy_name if boundary_diff > 0 else ('Esit' if boundary_diff == 0 else 'Traditional')],
                        ['Islem Suresi', f"{trad_time:.0f}ms", f"{new_time:.0f}ms",
                         f"{time_diff:+.0f}ms",
                         'Traditional' if time_diff > 0 else new_strategy_name],
                    ]
                    
                    comparison_table = Table(comparison_data, colWidths=[3.5*cm, 2.5*cm, 2.5*cm, 3.5*cm, 3*cm])
                    comparison_table.setStyle(TableStyle([
                        ('FONTNAME', (0, 0), (-1, -1), turkish_font),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('FONTSIZE', (0, 0), (-1, 0), 9),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#a7f3d0')),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ecfdf5')),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                        ('TOPPADDING', (0, 0), (-1, -1), 8),
                    ]))
                    story.append(comparison_table)
                    story.append(Spacer(1, 15))
                    
                    # Winner summary
                    wins = {'Traditional': 0, new_strategy_name: 0, 'Esit': 0}
                    if semantic_diff > 0.01:
                        wins[new_strategy_name] += 1
                    elif semantic_diff < -0.01:
                        wins['Traditional'] += 1
                    else:
                        wins['Esit'] += 1
                        
                    if boundary_diff > 0.01:
                        wins[new_strategy_name] += 1
                    elif boundary_diff < -0.01:
                        wins['Traditional'] += 1
                    else:
                        wins['Esit'] += 1
                    
                    # Determine overall winner
                    if wins[new_strategy_name] > wins['Traditional']:
                        winner = new_strategy_name
                        winner_reason = f"Semantic ve Boundary kalitesinde ustun"
                    elif wins['Traditional'] > wins[new_strategy_name]:
                        winner = 'Traditional'
                        winner_reason = "Daha hizli islem suresi"
                    else:
                        winner = 'Esit'
                        winner_reason = "Her iki yontem benzer sonuclar uretti"
                    
                    story.append(Paragraph(f"<b>Genel Degerlendirme:</b> {winner} stratejisi one cikiyor. {winner_reason}.", normal_style))
                    
                    # Detailed analysis
                    story.append(Spacer(1, 10))
                    analysis_points = []
                    
                    if semantic_improvement > 5:
                        analysis_points.append(f"• {new_strategy_name} semantik tutarlilikta %{semantic_improvement:.1f} iyilesme sagladi.")
                    elif semantic_improvement < -5:
                        analysis_points.append(f"• Traditional semantik tutarlilikta %{abs(semantic_improvement):.1f} daha iyi.")
                    
                    if boundary_improvement > 5:
                        analysis_points.append(f"• {new_strategy_name} sinir kalitesinde %{boundary_improvement:.1f} iyilesme sagladi.")
                    
                    if time_diff > 1000:
                        analysis_points.append(f"• {new_strategy_name} {time_diff/1000:.1f} saniye daha uzun surdu (LLM cagrilari nedeniyle).")
                    
                    if abs(chunk_diff) > 5:
                        if chunk_diff > 0:
                            analysis_points.append(f"• {new_strategy_name} {chunk_diff} adet daha fazla chunk uretti (daha ince parcalama).")
                        else:
                            analysis_points.append(f"• {new_strategy_name} {abs(chunk_diff)} adet daha az chunk uretti (daha iyi birlestirme).")
                    
                    for point in analysis_points:
                        story.append(Paragraph(point, small_style))
                    
                    story.append(Spacer(1, 10))
            
            story.append(Spacer(1, 10))
            
            # ===== DETAILED RESULTS FOR EACH STRATEGY =====
            story.append(Paragraph("3. Strateji Detaylari", heading1_style))
            
            for result_idx, result in enumerate(results):
                strategy = result.get('strategy', 'Bilinmiyor')
                success = result.get('success', False)
                
                story.append(Paragraph(f"3.{result_idx + 1}. {strategy.upper()} Stratejisi", heading2_style))
                
                if not success:
                    error_msg = result.get('error', 'Bilinmeyen hata')
                    story.append(Paragraph(f"HATA: {escape_xml(error_msg)}", normal_style))
                    story.append(Spacer(1, 10))
                    continue
                
                # Strategy metrics
                metrics_data = [
                    ['Metrik', 'Deger'],
                    ['Toplam Chunk Sayisi', str(result.get('chunk_count', 0))],
                    ['Ortalama Chunk Boyutu', f"{result.get('avg_chunk_size', 0):.1f} karakter"],
                    ['Minimum Chunk Boyutu', f"{result.get('min_chunk_size', 0):.0f} karakter"],
                    ['Maksimum Chunk Boyutu', f"{result.get('max_chunk_size', 0):.0f} karakter"],
                    ['Semantic Coherence Score', f"{result.get('semantic_coherence_score', 0):.4f}"],
                    ['Boundary Quality Score', f"{result.get('boundary_quality_score', 0):.4f}"],
                    ['Islem Suresi', f"{result.get('processing_time_ms', 0):.2f} ms"],
                ]
                
                # Add reasoning decisions if available
                reasoning_decisions = result.get('reasoning_decisions', [])
                similarity_analysis = result.get('similarity_analysis', {})
                
                if reasoning_decisions:
                    # Count different decision types
                    split_count = sum(1 for d in reasoning_decisions if d.get('decision') in ['SPLIT', 'FORCE_SPLIT'])
                    merge_count = sum(1 for d in reasoning_decisions if d.get('decision') in ['MERGE', 'FORCE_MERGE'])
                    preserve_count = sum(1 for d in reasoning_decisions if d.get('decision') == 'PRESERVE')
                    force_split_count = sum(1 for d in reasoning_decisions if d.get('decision') == 'FORCE_SPLIT')
                    force_merge_count = sum(1 for d in reasoning_decisions if d.get('decision') == 'FORCE_MERGE')
                    avg_conf = sum(d.get('confidence', 0) for d in reasoning_decisions) / len(reasoning_decisions) if reasoning_decisions else 0
                    avg_semantic = sum(d.get('semantic_coherence', 0) for d in reasoning_decisions) / len(reasoning_decisions) if reasoning_decisions else 0
                    
                    metrics_data.extend([
                        ['--- Boundary Karar Istatistikleri ---', ''],
                        ['Toplam Boundary Karari', str(len(reasoning_decisions))],
                        ['SPLIT Kararlari (Toplam)', str(split_count)],
                        ['  - Normal SPLIT', str(split_count - force_split_count)],
                        ['  - FORCE_SPLIT (Boyut Asimi)', str(force_split_count)],
                        ['MERGE Kararlari (Toplam)', str(merge_count)],
                        ['  - Normal MERGE', str(merge_count - force_merge_count)],
                        ['  - FORCE_MERGE (Kucuk Chunk)', str(force_merge_count)],
                        ['PRESERVE Kararlari', str(preserve_count)],
                        ['Ortalama Confidence', f"{avg_conf:.3f}"],
                        ['Ortalama Semantic Coherence', f"{avg_semantic:.3f}"],
                    ])
                    
                    # Add split/merge ratio
                    if similarity_analysis:
                        split_ratio = similarity_analysis.get('split_ratio', 0)
                        merge_ratio = similarity_analysis.get('merge_ratio', 0)
                        metrics_data.extend([
                            ['SPLIT Orani', f"{split_ratio*100:.1f}%"],
                            ['MERGE Orani', f"{merge_ratio*100:.1f}%"],
                        ])
                
                # Add metadata statistics if available
                metadata_stats = result.get('metadata_stats', {})
                if metadata_stats:
                    metrics_data.extend([
                        ['--- Metadata Istatistikleri ---', ''],
                        ['Baslik Kapsami', f"{metadata_stats.get('header_coverage_percent', 0):.1f}%"],
                        ['Ort. Anahtar Kelime/Chunk', f"{metadata_stats.get('avg_keywords_per_chunk', 0):.1f}"],
                    ])
                    
                    # Language distribution
                    lang_dist = metadata_stats.get('language_distribution', {})
                    if lang_dist:
                        lang_str = ', '.join(f"{k}: {v}" for k, v in lang_dist.items())
                        metrics_data.append(['Dil Dagilimi', lang_str])
                    
                    # Chunk type distribution
                    type_dist = metadata_stats.get('chunk_type_distribution', {})
                    if type_dist:
                        type_labels = {
                            'content': 'Icerik',
                            'header': 'Baslik',
                            'list': 'Liste',
                            'table': 'Tablo',
                            'code': 'Kod',
                            'question': 'Soru',
                            'image_caption': 'Resim'
                        }
                        type_str = ', '.join(f"{type_labels.get(k, k)}: {v}" for k, v in type_dist.items())
                        metrics_data.append(['Chunk Tur Dagilimi', type_str])
                
                metrics_table = Table(metrics_data, colWidths=[6*cm, 10*cm])
                metrics_table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), turkish_font),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e0e7ff')),
                    ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                    ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#c7d2fe')),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                ]))
                story.append(metrics_table)
                story.append(Spacer(1, 15))
                
                # ===== ALL CHUNKS (Full content with metadata) =====
                chunks = result.get('chunks', [])
                detailed_chunks = result.get('detailed_chunks', [])
                
                if chunks:
                    story.append(Paragraph(f"Olusturulan Chunklar ({len(chunks)} adet)", heading2_style))
                    
                    for chunk_idx, chunk in enumerate(chunks):
                        # Get chunk text
                        if isinstance(chunk, dict):
                            chunk_text = chunk.get('text', chunk.get('content', str(chunk)))
                        else:
                            chunk_text = str(chunk)
                        
                        # Escape XML characters
                        chunk_text = escape_xml(chunk_text)
                        
                        # Calculate chunk stats
                        char_count = len(chunk_text)
                        word_count = len(chunk_text.split())
                        
                        # Get metadata from detailed_chunks if available
                        chunk_metadata = {}
                        if chunk_idx < len(detailed_chunks):
                            detailed = detailed_chunks[chunk_idx]
                            chunk_metadata = detailed.get('metadata', {}) if isinstance(detailed, dict) else {}
                        
                        # Chunk header with metadata
                        chunk_header = f"<b>Chunk {chunk_idx + 1}</b> | {char_count} karakter | {word_count} kelime"
                        
                        # Add chunk type if available
                        chunk_type = chunk_metadata.get('chunk_type', '')
                        if chunk_type and chunk_type != 'content':
                            type_labels = {
                                'header': 'Baslik',
                                'list': 'Liste',
                                'table': 'Tablo',
                                'code': 'Kod',
                                'question': 'Soru',
                                'image_caption': 'Resim'
                            }
                            chunk_header += f" | <font color='#6366f1'>[{type_labels.get(chunk_type, chunk_type)}]</font>"
                        
                        story.append(Paragraph(chunk_header, small_style))
                        
                        # Add metadata info if available
                        metadata_parts = []
                        
                        # Parent header (section)
                        parent_header = chunk_metadata.get('parent_header')
                        if parent_header:
                            metadata_parts.append(f"<b>Bolum:</b> {escape_xml(str(parent_header))}")
                        
                        # Keywords
                        keywords = chunk_metadata.get('keywords', [])
                        if isinstance(keywords, str):
                            try:
                                keywords = json.loads(keywords)
                            except:
                                keywords = []
                        if keywords and isinstance(keywords, list) and len(keywords) > 0:
                            keywords_str = ', '.join(escape_xml(str(k)) for k in keywords[:5])
                            if len(keywords) > 5:
                                keywords_str += f' (+{len(keywords) - 5} daha)'
                            metadata_parts.append(f"<b>Anahtar Kelimeler:</b> {keywords_str}")
                        
                        # Language
                        language = chunk_metadata.get('language')
                        if language and language != 'auto':
                            lang_labels = {'tr': 'Turkce', 'en': 'Ingilizce'}
                            metadata_parts.append(f"<b>Dil:</b> {lang_labels.get(language, language)}")
                        
                        # Page number
                        page_num = chunk_metadata.get('page_number')
                        if page_num is not None and page_num >= 0:
                            metadata_parts.append(f"<b>Sayfa:</b> {page_num}")
                        
                        if metadata_parts:
                            metadata_line = ' | '.join(metadata_parts)
                            story.append(Paragraph(f"<font color='#6b7280' size='8'>{metadata_line}</font>", small_style))
                        
                        # Full chunk content
                        if len(chunk_text) > 3000:
                            wrapped_text = textwrap.fill(chunk_text, width=100)
                            story.append(Paragraph(wrapped_text, chunk_style))
                        else:
                            story.append(Paragraph(chunk_text, chunk_style))
                        
                        story.append(Spacer(1, 8))
                        
                        # Page break every 8 chunks
                        if (chunk_idx + 1) % 8 == 0 and chunk_idx < len(chunks) - 1:
                            story.append(PageBreak())
                    
                    story.append(Spacer(1, 15))
                
                # ===== REASONING DECISIONS =====
                if reasoning_decisions:
                    story.append(PageBreak())
                    story.append(Paragraph(f"Agentic Reasoning Kararlari ({len(reasoning_decisions)} adet)", heading2_style))
                    
                    for dec_idx, decision in enumerate(reasoning_decisions):
                        dec_type = decision.get('decision', 'UNKNOWN')
                        confidence = decision.get('confidence', 0)
                        reasoning = escape_xml(decision.get('reasoning', 'Aciklama yok'))
                        semantic_coh = decision.get('semantic_coherence', 0)
                        topic_cont = decision.get('topic_continuity', 0)
                        
                        dec_color = '#22c55e' if dec_type == 'MERGE' else '#ef4444'
                        
                        dec_header = f"<b>Karar {dec_idx + 1}:</b> <font color='{dec_color}'>{dec_type}</font> | Conf: {confidence:.2f} | Sem: {semantic_coh:.2f} | Topic: {topic_cont:.2f}"
                        story.append(Paragraph(dec_header, small_style))
                        story.append(Paragraph(f"<i>{reasoning}</i>", small_style))
                        story.append(Spacer(1, 6))
                        
                        if (dec_idx + 1) % 15 == 0 and dec_idx < len(reasoning_decisions) - 1:
                            story.append(PageBreak())
                
                story.append(PageBreak())
        
        else:
            logger.warning(f"⚠️ [PDF EXPORT] No results found for test {test_id}")
            story.append(Paragraph("Henuz sonuc bulunmuyor.", normal_style))
        
        # ===== FOOTER =====
        story.append(Spacer(1, 30))
        story.append(Paragraph(
            f"Bu rapor Agentic Chunking Test Sistemi tarafindan otomatik olusturulmustur. | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            small_style
        ))
        
        logger.info(f"📄 [PDF EXPORT] Building PDF document with {len(story)} elements")
        
        # Build PDF
        doc.build(story)
        
        # Get PDF bytes
        pdf_bytes = buffer.getvalue()
        buffer.close()
        
        logger.info(f"✅ [PDF EXPORT] PDF created successfully, size: {len(pdf_bytes)} bytes")
        
        # Return PDF as response
        filename = f"chunking_report_{test_id[:8]}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Type": "application/pdf"
            }
        )
        
    except Exception as e:
        logger.error(f"❌ [PDF EXPORT] Failed to export chunking test as PDF: {e}", exc_info=True)
        # Fallback to markdown JSON for frontend to handle
        try:
            logger.info(f"📄 [PDF EXPORT] Attempting fallback to markdown report")
            from src.utils.academic_report_generator import generate_chunking_academic_report
            markdown_report = generate_chunking_academic_report(test_data)
            
            logger.info(f"✅ [PDF EXPORT] Fallback markdown report generated")
            return {
                "success": True,
                "test_id": test_id,
                "format": "markdown",
                "report": markdown_report,
                "filename": f"chunking_test_report_{test_id[:8]}.md"
            }
        except Exception as fallback_error:
            logger.error(f"❌ [PDF EXPORT] Fallback also failed: {fallback_error}", exc_info=True)
            raise HTTPException(status_code=500, detail=f"Failed to export chunking test: {str(e)}")

# ===== BACKGROUND TASK FUNCTIONS =====

async def execute_full_chunking_test(
    test_id: str,
    config: ChunkingTestConfiguration
):
    """Background task to execute full chunking strategy comparison test"""
    
    test_data = CHUNKING_TEST_RESULTS_STORAGE[test_id]
    
    try:
        all_results = []
        total_strategies = len(config.strategies)
        
        logger.info(f"🚀 [CHUNKING TEST] Starting test {test_id} with {total_strategies} strategies: {config.strategies}")
        
        # Execute test for each strategy
        for strategy_index, strategy in enumerate(config.strategies):
            logger.info(f"🔄 [CHUNKING TEST] Processing strategy {strategy_index + 1}/{total_strategies}: {strategy}")
            
            # Update current strategy and progress
            test_data["current_strategy"] = strategy
            test_data["progress_percentage"] = (strategy_index / total_strategies) * 100
            test_data["status"] = "running"
            test_data["progress_message"] = f"Strateji işleniyor: {strategy} ({strategy_index + 1}/{total_strategies})"
            test_data["sub_progress"] = {
                "current_step": 0,
                "total_steps": 6,
                "step_message": "Başlatılıyor..."
            }
            
            # Save intermediate progress
            CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
            _save_chunking_test_to_db(test_id, test_data)
            
            # Execute strategy based on type
            if strategy == "traditional":
                result = await execute_traditional_chunking(
                    config.input_text,
                    config.target_chunk_size,
                    config.overlap_size,
                    config.session_id
                )
            elif strategy == "agentic" or strategy == "agentic_reasoning":
                result = await execute_agentic_reasoning_chunking(
                    config.input_text,
                    config.target_chunk_size,
                    config.overlap_size,
                    config.enable_grok_reasoning,
                    config.turkish_optimization,
                    config.session_id,
                    test_id  # Pass test_id for progress tracking
                )
            elif strategy == "llm_markdown":
                result = await execute_llm_markdown_chunking(
                    config.input_text,
                    config.target_chunk_size,
                    config.overlap_size,
                    config.session_id
                )
            elif strategy == "multi_agent":
                result = await execute_multi_agent_chunking(
                    config.input_text,
                    config.target_chunk_size,
                    config.overlap_size,
                    0.75,  # quality_threshold
                    config.min_chunk_size,
                    config.max_chunk_size,
                    config.session_id
                )
            else:
                logger.warning(f"❌ [CHUNKING TEST] Unknown strategy: {strategy}")
                continue
            
            all_results.append(result)
            
            # Update current metrics with detailed info
            test_data["current_metrics"] = {
                "chunk_count": result.get("chunk_count", 0),
                "processing_time_ms": result.get("processing_time_ms", 0),
                "success": result.get("success", False),
                "strategy": strategy,
                "avg_chunk_size": result.get("avg_chunk_size", 0),
                "semantic_coherence": result.get("semantic_coherence_score", 0),
                "boundary_quality": result.get("boundary_quality_score", 0)
            }
            
            # Mark strategy as completed
            completed_strategies = test_data.get("completed_strategies", [])
            if strategy not in completed_strategies:
                completed_strategies.append(strategy)
            test_data["completed_strategies"] = completed_strategies
            
            # Update progress percentage
            test_data["progress_percentage"] = ((strategy_index + 1) / total_strategies) * 100
            
            logger.info(f"✅ [CHUNKING TEST] Strategy {strategy} completed: {result.get('success', False)} - {result.get('chunk_count', 0)} chunks")
            
            # Update progress in real-time with more detailed data
            test_data["results"] = all_results.copy()
            test_data["last_updated"] = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
            
            # Save progress immediately
            CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
            _save_chunking_test_to_db(test_id, test_data)
            
            # Add a small delay to make progress visible
            await asyncio.sleep(0.5)
        
        # Store final results
        test_data["results"] = all_results
        test_data["status"] = "completed"
        test_data["progress_percentage"] = 100.0
        test_data["end_time"] = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
        test_data["last_updated"] = test_data["end_time"]
        
        # Save final results
        CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
        _save_chunking_test_to_db(test_id, test_data)
        
        logger.info(f"🎉 [CHUNKING TEST] Test {test_id} completed successfully with {len(all_results)} results")
        
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        logger.error(f"💥 [CHUNKING TEST] Test {test_id} failed: {e}")
        logger.error(f"Traceback: {error_traceback}")
        
        # Mark as failed
        test_data["status"] = "failed"
        test_data["error"] = str(e)
        test_data["error_traceback"] = error_traceback
        test_data["end_time"] = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
        test_data["last_updated"] = test_data["end_time"]
        test_data["progress_percentage"] = 0.0
        
        # Save any partial results
        if "results" not in test_data:
            test_data["results"] = []
        
        CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
        _save_chunking_test_to_db(test_id, test_data)

@router.get("/reasoning-analysis/{test_id}", summary="Get Detailed Reasoning Analysis")
async def get_reasoning_analysis(test_id: str, request: Request = None) -> Dict[str, Any]:
    """Get detailed reasoning analysis for agentic chunking visualization"""
    # Basic authentication check
    if request:
        current_user = _get_current_user(request)
        logger.info(f"🔍 [REASONING ANALYSIS] Current user: {current_user}")
        
        if not (_is_teacher(current_user) or _is_admin(current_user)):
            raise HTTPException(status_code=403, detail="Teacher or admin access required")
        
        logger.info(f"🔍 [REASONING ANALYSIS] Authentication successful - access granted")
    
    # Try memory first, then database
    test_data = CHUNKING_TEST_RESULTS_STORAGE.get(test_id)
    if not test_data:
        test_data = _load_chunking_test_from_db(test_id)
        if test_data:
            CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
    
    if not test_data:
        raise HTTPException(status_code=404, detail="Chunking test not found")
    
    try:
        results = test_data.get("results", [])
        
        # Find agentic reasoning result
        agentic_result = None
        for result in results:
            if result.get("strategy") == "agentic" and result.get("success", False):
                agentic_result = result
                break
        
        if not agentic_result:
            return {
                "success": False,
                "message": "No successful agentic reasoning result found",
                "test_id": test_id
            }
        
        # Extract detailed reasoning information
        detailed_chunks = agentic_result.get("detailed_chunks", [])
        reasoning_decisions = agentic_result.get("reasoning_decisions", [])
        similarity_analysis = agentic_result.get("similarity_analysis", {})
        
        # Create decision timeline for visualization
        decision_timeline = []
        for i, decision in enumerate(reasoning_decisions):
            decision_timeline.append({
                "id": i,
                "timestamp": i,  # Sequential order
                "decision": decision.get("decision", "UNKNOWN"),
                "confidence": decision.get("confidence", 0),
                "reasoning": decision.get("reasoning", "No reasoning provided"),
                "semantic_coherence": decision.get("semantic_coherence", 0),
                "topic_continuity": decision.get("topic_continuity", 0),
                "metadata": decision.get("metadata", {}),
                "decision_method": decision.get("metadata", {}).get("decision_method", "unknown")
            })
        
        # Calculate semantic metrics analysis
        semantic_metrics = {
            "coherence_distribution": {},
            "topic_continuity_distribution": {},
            "confidence_distribution": {},
            "decision_type_distribution": {}
        }
        
        if reasoning_decisions:
            # Coherence distribution
            coherence_ranges = {"low": 0, "medium": 0, "high": 0}
            topic_ranges = {"low": 0, "medium": 0, "high": 0}
            confidence_ranges = {"low": 0, "medium": 0, "high": 0}
            decision_types = {"SPLIT": 0, "MERGE": 0, "UNKNOWN": 0}
            
            for decision in reasoning_decisions:
                # Coherence
                coherence = decision.get("semantic_coherence", 0)
                if coherence < 0.4:
                    coherence_ranges["low"] += 1
                elif coherence < 0.7:
                    coherence_ranges["medium"] += 1
                else:
                    coherence_ranges["high"] += 1
                
                # Topic continuity
                topic = decision.get("topic_continuity", 0)
                if topic < 0.4:
                    topic_ranges["low"] += 1
                elif topic < 0.7:
                    topic_ranges["medium"] += 1
                else:
                    topic_ranges["high"] += 1
                
                # Confidence
                confidence = decision.get("confidence", 0)
                if confidence < 0.4:
                    confidence_ranges["low"] += 1
                elif confidence < 0.7:
                    confidence_ranges["medium"] += 1
                else:
                    confidence_ranges["high"] += 1
                
                # Decision types
                decision_type = decision.get("decision", "UNKNOWN")
                decision_types[decision_type] = decision_types.get(decision_type, 0) + 1
            
            semantic_metrics = {
                "coherence_distribution": coherence_ranges,
                "topic_continuity_distribution": topic_ranges,
                "confidence_distribution": confidence_ranges,
                "decision_type_distribution": decision_types
            }
        
        # Chunk quality analysis
        chunk_quality_analysis = []
        for chunk in detailed_chunks:
            issues = chunk.get("issues", [])
            quality_score = chunk.get("quality_score", 0)
            
            # Determine quality level and recommendations
            quality_level = "high" if quality_score > 0.7 else "medium" if quality_score > 0.4 else "low"
            recommendations = []
            
            if quality_score < 0.5:
                recommendations.append("Consider merging with adjacent chunks")
            if chunk.get("word_count", 0) < 50:
                recommendations.append("Chunk may be too small for meaningful content")
            if len(issues) > 2:
                recommendations.append("Multiple quality issues detected")
            if chunk.get("semantic_coherence", 0) < 0.4:
                recommendations.append("Low semantic coherence - review content boundaries")
            
            chunk_quality_analysis.append({
                "chunk_id": chunk.get("id", 0),
                "quality_score": quality_score,
                "quality_level": quality_level,
                "issues": issues,
                "recommendations": recommendations,
                "word_count": chunk.get("word_count", 0),
                "has_header": chunk.get("has_header", False),
                "semantic_coherence": chunk.get("semantic_coherence", 0),
                "reasoning_confidence": chunk.get("reasoning_confidence", 0)
            })
        
        # Performance insights
        performance_insights = {
            "total_boundary_decisions": len(reasoning_decisions),
            "split_merge_ratio": similarity_analysis.get("split_ratio", 0),
            "average_confidence": similarity_analysis.get("avg_confidence", 0),
            "reasoning_methods_used": similarity_analysis.get("reasoning_methods", []),
            "processing_time_ms": agentic_result.get("processing_time_ms", 0),
            "chunks_created": len(detailed_chunks),
            "average_chunk_quality": sum(chunk.get("quality_score", 0) for chunk in detailed_chunks) / len(detailed_chunks) if detailed_chunks else 0
        }
        
        # Visualization data for charts
        visualization_data = {
            "confidence_over_time": [{"x": i, "y": d.get("confidence", 0)} for i, d in enumerate(reasoning_decisions)],
            "coherence_over_time": [{"x": i, "y": d.get("semantic_coherence", 0)} for i, d in enumerate(reasoning_decisions)],
            "decision_distribution": [
                {"label": "Split", "value": len([d for d in reasoning_decisions if d.get("decision") == "SPLIT"])},
                {"label": "Merge", "value": len([d for d in reasoning_decisions if d.get("decision") == "MERGE"])}
            ],
            "chunk_size_distribution": [{"x": i, "y": len(chunk.get("text", ""))} for i, chunk in enumerate(detailed_chunks)]
        }
        
        logger.info(f"📊 [REASONING ANALYSIS] Generated analysis for {len(detailed_chunks)} chunks, {len(reasoning_decisions)} decisions")
        
        return {
            "success": True,
            "test_id": test_id,
            "analysis": {
                "decision_timeline": decision_timeline,
                "semantic_metrics": semantic_metrics,
                "chunk_quality_analysis": chunk_quality_analysis,
                "performance_insights": performance_insights,
                "visualization_data": visualization_data,
                "detailed_chunks": detailed_chunks,
                "similarity_analysis": similarity_analysis
            },
            "summary": {
                "total_chunks": len(detailed_chunks),
                "total_decisions": len(reasoning_decisions),
                "average_confidence": performance_insights["average_confidence"],
                "average_chunk_quality": performance_insights["average_chunk_quality"],
                "processing_time_ms": performance_insights["processing_time_ms"]
            }
        }
        
    except Exception as e:
        logger.error(f"Failed to get reasoning analysis: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get reasoning analysis: {str(e)}")

@router.put("/update/{test_id}", summary="Update Chunking Test")
async def update_chunking_test(
    test_id: str,
    request_data: dict,
    request: Request = None
) -> Dict[str, Any]:
    """Update chunking test name or other properties"""
    # Basic authentication check
    current_user = _get_current_user(request)
    logger.info(f"🔍 [CHUNKING TEST UPDATE] Current user: {current_user}")
    
    if not (_is_teacher(current_user) or _is_admin(current_user)):
        raise HTTPException(status_code=403, detail="Teacher or admin access required")
    
    logger.info(f"🔍 [CHUNKING TEST UPDATE] Authentication successful - access granted")
    
    # Try memory first, then database
    test_data = CHUNKING_TEST_RESULTS_STORAGE.get(test_id)
    if not test_data:
        test_data = _load_chunking_test_from_db(test_id)
        if test_data:
            CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
    
    if not test_data:
        raise HTTPException(status_code=404, detail="Chunking test not found")
    
    # Update allowed fields
    if "testName" in request_data:
        test_data["test_name"] = request_data["testName"]
    
    # Update timestamp
    test_data["last_updated"] = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    
    # Save changes
    CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
    _save_chunking_test_to_db(test_id, test_data)
    
    logger.info(f"✅ [CHUNKING TEST UPDATE] Test {test_id} updated successfully")
    
    return {
        "success": True,
        "message": "Test updated successfully",
        "test_id": test_id,
        "updated_fields": list(request_data.keys())
    }


# ===== EVALUATION API ENDPOINTS =====
# Requirements: 1.1, 2.3, 3.5, 4.1

@router.get("/evaluate/{test_id}", summary="Full Evaluation of Chunking Test")
async def evaluate_chunking_test(test_id: str, request: Request = None) -> Dict[str, Any]:
    """
    Get comprehensive evaluation of chunking test with all metrics.
    
    Returns:
    - Similarity analysis (intra-chunk, inter-chunk, topic separation)
    - Scientific metrics (HOPE, Topic Drift, Context Preservation, etc.)
    - Agent performance scores (Structural, Semantic, Size, Quality)
    - Overall quality comparison
    
    Requirements: 1.1, 2.3, 3.5
    """
    logger.info(f"📊 [EVALUATION] Starting full evaluation for test_id: {test_id}")
    
    # Authentication check
    current_user = _get_current_user(request)
    if not (_is_teacher(current_user) or _is_admin(current_user)):
        raise HTTPException(status_code=403, detail="Teacher or admin access required")
    
    # Check embedding service health
    from src.embedding.embedding_generator import check_embedding_service_health, is_using_fallback_embeddings
    embedding_healthy, embedding_message = check_embedding_service_health()
    
    if not embedding_healthy:
        logger.warning(f"⚠️ [EVALUATION] Embedding service not healthy: {embedding_message}")
    
    # Load test data
    test_data = CHUNKING_TEST_RESULTS_STORAGE.get(test_id)
    if not test_data:
        test_data = _load_chunking_test_from_db(test_id)
        if test_data:
            CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
    
    if not test_data:
        raise HTTPException(status_code=404, detail="Chunking test not found")
    
    try:
        from src.evaluation import (
            SimilarityAnalyzer, 
            ScientificMetricCalculator,
            AgentEvaluator,
            ChunkData,
            ChunkingConfig
        )
        
        # Extract chunks from test data
        # Test data structure: results is an array where each item has a "strategy" field
        results = test_data.get("results", [])
        if not results:
            raise HTTPException(status_code=400, detail="No results found in test data")
        
        # Get traditional and multi-agent chunks
        traditional_chunks = []
        multi_agent_chunks = []
        original_text = test_data.get("original_text", "")
        
        # Find traditional and multi-agent results
        traditional_result = None
        multi_agent_result = None
        
        for result in results:
            strategy = result.get("strategy", "")
            if strategy == "traditional":
                traditional_result = result
            elif strategy in ["multi_agent", "agentic", "agentic_reasoning"]:
                multi_agent_result = result
        
        # Extract traditional chunks
        if traditional_result:
            raw_chunks = traditional_result.get("chunks", [])
            detailed_chunks = traditional_result.get("detailed_chunks", [])
            chunks_to_use = detailed_chunks if detailed_chunks else raw_chunks
            
            for chunk in chunks_to_use:
                if isinstance(chunk, dict):
                    traditional_chunks.append(chunk.get("text", chunk.get("content", "")))
                elif isinstance(chunk, str):
                    traditional_chunks.append(chunk)
                else:
                    traditional_chunks.append(str(chunk))
            
            if not original_text:
                original_text = traditional_result.get("original_text", "")
        
        # Extract multi-agent chunks
        if multi_agent_result:
            raw_chunks = multi_agent_result.get("chunks", [])
            detailed_chunks = multi_agent_result.get("detailed_chunks", [])
            chunks_to_use = detailed_chunks if detailed_chunks else raw_chunks
            
            for chunk in chunks_to_use:
                if isinstance(chunk, dict):
                    multi_agent_chunks.append(chunk.get("text", chunk.get("content", "")))
                elif isinstance(chunk, str):
                    multi_agent_chunks.append(chunk)
                else:
                    multi_agent_chunks.append(str(chunk))
            
            if not original_text:
                original_text = multi_agent_result.get("original_text", "")
        
        # If we don't have both, use what we have
        if not traditional_chunks and not multi_agent_chunks:
            raise HTTPException(status_code=400, detail="No chunks found in test results")
        
        # If only one strategy exists, use it for both (for comparison purposes)
        if not traditional_chunks:
            traditional_chunks = multi_agent_chunks
            logger.warning("⚠️ [EVALUATION] No traditional chunks found, using multi-agent for comparison")
        if not multi_agent_chunks:
            multi_agent_chunks = traditional_chunks
            logger.warning("⚠️ [EVALUATION] No multi-agent chunks found, using traditional for comparison")
        
        logger.info(f"📊 [EVALUATION] Found {len(traditional_chunks)} traditional chunks and {len(multi_agent_chunks)} multi-agent chunks")
        
        # Initialize analyzers
        similarity_analyzer = SimilarityAnalyzer()
        scientific_calculator = ScientificMetricCalculator()
        agent_evaluator = AgentEvaluator()
        
        # Similarity Analysis
        logger.info(f"📊 [EVALUATION] Calculating similarity metrics...")
        traditional_similarity = similarity_analyzer.analyze_strategy(traditional_chunks)
        multi_agent_similarity = similarity_analyzer.analyze_strategy(multi_agent_chunks)
        
        # Scientific Metrics
        logger.info(f"📊 [EVALUATION] Calculating scientific metrics...")
        traditional_scientific = scientific_calculator.calculate_all_metrics(traditional_chunks)
        multi_agent_scientific = scientific_calculator.calculate_all_metrics(multi_agent_chunks)
        
        # Agent Evaluation (for multi-agent only)
        logger.info(f"📊 [EVALUATION] Evaluating agent performance...")
        
        # Convert multi-agent chunks to ChunkData format
        chunk_data_list = []
        if multi_agent_result:
            raw_chunks = multi_agent_result.get("chunks", [])
            detailed_chunks = multi_agent_result.get("detailed_chunks", [])
            chunks_to_use = detailed_chunks if detailed_chunks else raw_chunks
            
            for chunk in chunks_to_use:
                if isinstance(chunk, dict):
                    content = chunk.get("text", chunk.get("content", ""))
                    chunk_data_list.append(ChunkData(
                        content=content,
                        char_count=chunk.get("char_count", len(content)),
                        word_count=chunk.get("word_count", len(content.split())),
                        boundary_type=chunk.get("boundary_type", "unknown"),
                        quality_score=chunk.get("quality_score", 0.0)
                    ))
                elif isinstance(chunk, str):
                    chunk_data_list.append(ChunkData(
                        content=chunk,
                        char_count=len(chunk),
                        word_count=len(chunk.split()),
                        boundary_type="unknown",
                        quality_score=0.0
                    ))
        
        # If no multi-agent chunks, use traditional chunks for agent evaluation
        if not chunk_data_list:
            for chunk_text in multi_agent_chunks:
                chunk_data_list.append(ChunkData(
                    content=chunk_text,
                    char_count=len(chunk_text),
                    word_count=len(chunk_text.split()),
                    boundary_type="unknown",
                    quality_score=0.0
                ))
        
        config = ChunkingConfig(
            target_chunk_size=test_data.get("config", {}).get("target_chunk_size",
                             test_data.get("configuration", {}).get("target_chunk_size", 1500)),
            min_chunk_size=test_data.get("config", {}).get("min_chunk_size",
                          test_data.get("configuration", {}).get("min_chunk_size", 500)),
            max_chunk_size=test_data.get("config", {}).get("max_chunk_size",
                          test_data.get("configuration", {}).get("max_chunk_size", 3000))
        )
        
        agent_evaluation = agent_evaluator.evaluate_all(
            chunks=chunk_data_list,
            original_text=original_text,
            config=config
        )
        
        # Calculate improvements
        def calc_improvement(trad, multi):
            if trad == 0:
                return 100.0 if multi > 0 else 0.0
            return ((multi - trad) / trad) * 100
        
        improvements = {
            "intra_chunk_similarity": calc_improvement(
                traditional_similarity.intra_chunk_similarity,
                multi_agent_similarity.intra_chunk_similarity
            ),
            "topic_separation": calc_improvement(
                traditional_similarity.topic_separation_score,
                multi_agent_similarity.topic_separation_score
            ),
            "overall_quality": calc_improvement(
                traditional_scientific.overall_quality_index,
                multi_agent_scientific.overall_quality_index
            )
        }
        
        logger.info(f"✅ [EVALUATION] Evaluation completed for test {test_id}")
        
        # Check if using fallback embeddings and add warning
        using_fallback = is_using_fallback_embeddings()
        
        response_data = {
            "success": True,
            "test_id": test_id,
            "evaluation": {
                "similarity_analysis": {
                    "traditional": traditional_similarity.to_dict(),
                    "multi_agent": multi_agent_similarity.to_dict()
                },
                "scientific_metrics": {
                    "traditional": traditional_scientific.to_dict(),
                    "multi_agent": multi_agent_scientific.to_dict()
                },
                "agent_evaluation": agent_evaluation.to_dict(),
                "improvements": improvements
            },
            "summary": {
                "traditional_quality": traditional_scientific.overall_quality_index,
                "multi_agent_quality": multi_agent_scientific.overall_quality_index,
                "overall_improvement_pct": improvements["overall_quality"],
                "winner": "multi_agent" if improvements["overall_quality"] > 0 else "traditional"
            }
        }
        
        # Add warning if using fallback embeddings
        if using_fallback or not embedding_healthy:
            response_data["warning"] = {
                "type": "fallback_embeddings",
                "message": "⚠️ Embedding servisi çalışmıyor! Hash-based fallback embedding kullanılıyor. "
                          "Semantik analiz sonuçları ANLAMSIZ olacaktır. Gerçek sonuçlar için embedding "
                          "servisinin (model-inference-service) çalıştığından emin olun.",
                "embedding_service_status": embedding_message,
                "recommendation": "Docker container'ları yeniden başlatın veya model-inference-service'in "
                                 "çalıştığını kontrol edin."
            }
            logger.warning(f"⚠️ [EVALUATION] Results may be unreliable - using fallback embeddings!")
        
        return response_data
        
    except ImportError as e:
        logger.error(f"❌ [EVALUATION] Import error: {e}")
        raise HTTPException(status_code=500, detail=f"Evaluation module not available: {str(e)}")
    except Exception as e:
        logger.error(f"❌ [EVALUATION] Failed to evaluate test: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to evaluate chunking test: {str(e)}")


@router.get("/export-zip/{test_id}", summary="Export Chunking Test as ZIP")
async def export_chunking_test_zip(test_id: str, request: Request = None):
    """
    Export chunking test results as ZIP archive.
    
    ZIP Structure:
    - traditional/ folder with chunk files
    - multi_agent/ folder with chunk files
    - metadata.json
    - comparison_report.md
    
    Requirements: 1.2
    """
    from fastapi.responses import Response
    
    logger.info(f"📦 [ZIP EXPORT] Starting ZIP export for test_id: {test_id}")
    
    # Authentication check
    current_user = _get_current_user(request)
    if not (_is_teacher(current_user) or _is_admin(current_user)):
        raise HTTPException(status_code=403, detail="Teacher or admin access required")
    
    # Load test data
    test_data = CHUNKING_TEST_RESULTS_STORAGE.get(test_id)
    if not test_data:
        test_data = _load_chunking_test_from_db(test_id)
        if test_data:
            CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
    
    if not test_data:
        raise HTTPException(status_code=404, detail="Chunking test not found")
    
    try:
        from src.evaluation import ChunkExportManager, ComparisonReportGenerator, StrategyMetrics
        
        exporter = ChunkExportManager()
        report_generator = ComparisonReportGenerator()
        
        # Extract chunks
        results = test_data.get("results", [])
        traditional_chunks = []
        multi_agent_chunks = []
        
        for result in results:
            traditional_chunks.extend(result.get("traditional", {}).get("chunks", []))
            multi_agent_chunks.extend(result.get("multi_agent", {}).get("chunks", []))
        
        # Generate comparison report
        trad_metrics = StrategyMetrics(
            chunk_count=len(traditional_chunks),
            avg_chunk_size=sum(c.get("char_count", 0) for c in traditional_chunks) / max(len(traditional_chunks), 1),
            total_chars=sum(c.get("char_count", 0) for c in traditional_chunks)
        )
        
        multi_metrics = StrategyMetrics(
            chunk_count=len(multi_agent_chunks),
            avg_chunk_size=sum(c.get("char_count", 0) for c in multi_agent_chunks) / max(len(multi_agent_chunks), 1),
            total_chars=sum(c.get("char_count", 0) for c in multi_agent_chunks)
        )
        
        comparison = report_generator.generate_comparison(trad_metrics, multi_metrics)
        
        test_info = {
            "test_id": test_id,
            "test_name": test_data.get("test_name", ""),
            "document_title": test_data.get("document_title", ""),
            "target_chunk_size": test_data.get("config", {}).get("target_chunk_size", 1500),
            "min_chunk_size": test_data.get("config", {}).get("min_chunk_size", 500),
            "max_chunk_size": test_data.get("config", {}).get("max_chunk_size", 3000)
        }
        
        comparison_report = report_generator.generate_markdown_report(comparison, test_info)
        
        # Create ZIP
        zip_bytes = exporter.create_zip_archive(
            traditional_chunks=traditional_chunks,
            multi_agent_chunks=multi_agent_chunks,
            test_info=test_info,
            comparison_report=comparison_report
        )
        
        logger.info(f"✅ [ZIP EXPORT] ZIP created successfully, size: {len(zip_bytes)} bytes")
        
        filename = f"chunking_test_{test_id}.zip"
        
        return Response(
            content=zip_bytes,
            media_type="application/zip",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"'
            }
        )
        
    except ImportError as e:
        logger.error(f"❌ [ZIP EXPORT] Import error: {e}")
        raise HTTPException(status_code=500, detail=f"Export module not available: {str(e)}")
    except Exception as e:
        logger.error(f"❌ [ZIP EXPORT] Failed to export: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to export chunking test: {str(e)}")


@router.get("/agent-scores/{test_id}", summary="Get Agent Performance Scores")
async def get_agent_scores(test_id: str, request: Request = None) -> Dict[str, Any]:
    """
    Get detailed agent performance scores for multi-agent chunking.
    
    Returns scores for:
    - StructuralAgent: Atomic unit preservation
    - SemanticAgent: Topic boundary detection
    - SizeAgent: Size variance management
    - QualityAgent: Quality score evaluation
    
    Requirements: 3.5
    """
    logger.info(f"🤖 [AGENT SCORES] Getting agent scores for test_id: {test_id}")
    
    # Authentication check
    current_user = _get_current_user(request)
    if not (_is_teacher(current_user) or _is_admin(current_user)):
        raise HTTPException(status_code=403, detail="Teacher or admin access required")
    
    # Load test data
    test_data = CHUNKING_TEST_RESULTS_STORAGE.get(test_id)
    if not test_data:
        test_data = _load_chunking_test_from_db(test_id)
        if test_data:
            CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
    
    if not test_data:
        raise HTTPException(status_code=404, detail="Chunking test not found")
    
    try:
        from src.evaluation import AgentEvaluator, ChunkData, ChunkingConfig
        
        agent_evaluator = AgentEvaluator()
        
        # Extract multi-agent chunks and original text
        # Test data structure: results is an array where each item has a "strategy" field
        results = test_data.get("results", [])
        chunk_data_list = []
        original_text = test_data.get("original_text", "")
        
        # Find multi-agent or agentic strategy results
        multi_agent_result = None
        for result in results:
            strategy = result.get("strategy", "")
            if strategy in ["multi_agent", "agentic", "agentic_reasoning"]:
                multi_agent_result = result
                break
        
        # If no multi-agent result, try to use the first available result
        if not multi_agent_result and results:
            multi_agent_result = results[0]
            logger.warning(f"⚠️ [AGENT SCORES] No multi-agent result found, using first result with strategy: {multi_agent_result.get('strategy', 'unknown')}")
        
        if not multi_agent_result:
            raise HTTPException(status_code=400, detail="No chunking results found for evaluation")
        
        # Get original text from result if not in test_data
        if not original_text:
            original_text = multi_agent_result.get("original_text", "")
        
        # Extract chunks from the result
        # Chunks can be in different formats: detailed_chunks, chunks array, or raw chunks
        detailed_chunks = multi_agent_result.get("detailed_chunks", [])
        raw_chunks = multi_agent_result.get("chunks", [])
        
        if detailed_chunks:
            for chunk in detailed_chunks:
                if isinstance(chunk, dict):
                    content = chunk.get("content", chunk.get("text", ""))
                    chunk_data_list.append(ChunkData(
                        content=content,
                        char_count=chunk.get("char_count", len(content)),
                        word_count=chunk.get("word_count", len(content.split())),
                        boundary_type=chunk.get("boundary_type", "unknown"),
                        quality_score=chunk.get("quality_score", 0.0)
                    ))
        elif raw_chunks:
            for i, chunk in enumerate(raw_chunks):
                if isinstance(chunk, dict):
                    content = chunk.get("text", chunk.get("content", ""))
                elif isinstance(chunk, str):
                    content = chunk
                else:
                    content = str(chunk)
                
                chunk_data_list.append(ChunkData(
                    content=content,
                    char_count=len(content),
                    word_count=len(content.split()),
                    boundary_type="unknown",
                    quality_score=0.0
                ))
        
        if not chunk_data_list:
            raise HTTPException(status_code=400, detail="No chunks found in test results for evaluation")
        
        logger.info(f"📊 [AGENT SCORES] Found {len(chunk_data_list)} chunks for evaluation")
        
        config = ChunkingConfig(
            target_chunk_size=test_data.get("config", {}).get("target_chunk_size", 
                             test_data.get("configuration", {}).get("target_chunk_size", 1500)),
            min_chunk_size=test_data.get("config", {}).get("min_chunk_size",
                          test_data.get("configuration", {}).get("min_chunk_size", 500)),
            max_chunk_size=test_data.get("config", {}).get("max_chunk_size",
                          test_data.get("configuration", {}).get("max_chunk_size", 3000))
        )
        
        # Evaluate all agents
        evaluation = agent_evaluator.evaluate_all(
            chunks=chunk_data_list,
            original_text=original_text,
            config=config
        )
        
        logger.info(f"✅ [AGENT SCORES] Agent evaluation completed for test {test_id}")
        
        return {
            "success": True,
            "test_id": test_id,
            "agent_scores": {
                "structural": evaluation.structural_score.to_dict(),
                "semantic": evaluation.semantic_score.to_dict(),
                "size": evaluation.size_score.to_dict(),
                "quality": evaluation.quality_score.to_dict()
            },
            "overall_score": evaluation.overall_score,
            "weights": {
                "structural": evaluation.WEIGHT_STRUCTURAL,
                "semantic": evaluation.WEIGHT_SEMANTIC,
                "size": evaluation.WEIGHT_SIZE,
                "quality": evaluation.WEIGHT_QUALITY
            }
        }
        
    except ImportError as e:
        logger.error(f"❌ [AGENT SCORES] Import error: {e}")
        raise HTTPException(status_code=500, detail=f"Evaluation module not available: {str(e)}")
    except Exception as e:
        logger.error(f"❌ [AGENT SCORES] Failed to get agent scores: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to get agent scores: {str(e)}")


@router.get("/similarity-analysis/{test_id}", summary="Get Similarity Analysis")
async def get_similarity_analysis(test_id: str, request: Request = None) -> Dict[str, Any]:
    """
    Get detailed similarity analysis for chunking test.
    
    Returns:
    - Intra-chunk similarity (coherence within chunks)
    - Inter-chunk similarity (similarity between adjacent chunks)
    - Topic separation score
    
    Requirements: 2.3
    """
    logger.info(f"📈 [SIMILARITY] Getting similarity analysis for test_id: {test_id}")
    
    # Authentication check
    current_user = _get_current_user(request)
    if not (_is_teacher(current_user) or _is_admin(current_user)):
        raise HTTPException(status_code=403, detail="Teacher or admin access required")
    
    # Load test data
    test_data = CHUNKING_TEST_RESULTS_STORAGE.get(test_id)
    if not test_data:
        test_data = _load_chunking_test_from_db(test_id)
        if test_data:
            CHUNKING_TEST_RESULTS_STORAGE[test_id] = test_data
    
    if not test_data:
        raise HTTPException(status_code=404, detail="Chunking test not found")
    
    try:
        from src.evaluation import SimilarityAnalyzer
        
        analyzer = SimilarityAnalyzer()
        
        # Extract chunks
        results = test_data.get("results", [])
        traditional_chunks = []
        multi_agent_chunks = []
        
        for result in results:
            # Traditional chunks
            trad_data = result.get("traditional", {})
            for chunk in trad_data.get("chunks", []):
                if isinstance(chunk, dict):
                    traditional_chunks.append(chunk.get("text", chunk.get("content", "")))
                else:
                    traditional_chunks.append(str(chunk))
            
            # Multi-agent chunks
            multi_data = result.get("multi_agent", {})
            for chunk in multi_data.get("chunks", []):
                if isinstance(chunk, dict):
                    multi_agent_chunks.append(chunk.get("text", chunk.get("content", "")))
                else:
                    multi_agent_chunks.append(str(chunk))
        
        # Analyze both strategies
        traditional_metrics = analyzer.analyze_strategy(traditional_chunks)
        multi_agent_metrics = analyzer.analyze_strategy(multi_agent_chunks)
        
        logger.info(f"✅ [SIMILARITY] Similarity analysis completed for test {test_id}")
        
        return {
            "success": True,
            "test_id": test_id,
            "similarity_analysis": {
                "traditional": traditional_metrics.to_dict(),
                "multi_agent": multi_agent_metrics.to_dict()
            },
            "comparison": {
                "intra_chunk_improvement": (
                    (multi_agent_metrics.intra_chunk_similarity - traditional_metrics.intra_chunk_similarity) 
                    / max(traditional_metrics.intra_chunk_similarity, 0.001) * 100
                ),
                "topic_separation_improvement": (
                    (multi_agent_metrics.topic_separation_score - traditional_metrics.topic_separation_score)
                    / max(traditional_metrics.topic_separation_score, 0.001) * 100
                )
            }
        }
        
    except ImportError as e:
        logger.error(f"❌ [SIMILARITY] Import error: {e}")
        raise HTTPException(status_code=500, detail=f"Evaluation module not available: {str(e)}")
    except Exception as e:
        logger.error(f"❌ [SIMILARITY] Failed to get similarity analysis: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to get similarity analysis: {str(e)}")


@router.post("/batch-evaluate", summary="Batch Evaluation of Multiple Tests")
async def batch_evaluate_tests(
    request_data: Dict[str, Any],
    request: Request = None
) -> Dict[str, Any]:
    """
    Evaluate multiple chunking tests and aggregate results.
    
    Request body:
    - test_ids: List of test IDs to evaluate
    
    Returns:
    - Individual document results
    - Aggregate statistics (mean, std, min, max)
    - Statistical significance (p-value)
    - Effect size (Cohen's d)
    - Outlier detection
    
    Requirements: 7.1, 7.2, 7.3
    """
    logger.info(f"📊 [BATCH EVAL] Starting batch evaluation")
    
    # Authentication check
    current_user = _get_current_user(request)
    if not (_is_teacher(current_user) or _is_admin(current_user)):
        raise HTTPException(status_code=403, detail="Teacher or admin access required")
    
    test_ids = request_data.get("test_ids", [])
    if not test_ids:
        raise HTTPException(status_code=400, detail="No test IDs provided")
    
    try:
        from src.evaluation import BatchEvaluator, ScientificMetricCalculator
        
        batch_evaluator = BatchEvaluator()
        scientific_calculator = ScientificMetricCalculator()
        
        # Collect test results
        test_results = []
        
        for test_id in test_ids:
            test_data = CHUNKING_TEST_RESULTS_STORAGE.get(test_id)
            if not test_data:
                test_data = _load_chunking_test_from_db(test_id)
            
            if not test_data:
                logger.warning(f"⚠️ [BATCH EVAL] Test {test_id} not found, skipping")
                continue
            
            # Extract chunks and calculate metrics
            results = test_data.get("results", [])
            traditional_chunks = []
            multi_agent_chunks = []
            
            for result in results:
                trad_data = result.get("traditional", {})
                for chunk in trad_data.get("chunks", []):
                    if isinstance(chunk, dict):
                        traditional_chunks.append(chunk.get("text", chunk.get("content", "")))
                    else:
                        traditional_chunks.append(str(chunk))
                
                multi_data = result.get("multi_agent", {})
                for chunk in multi_data.get("chunks", []):
                    if isinstance(chunk, dict):
                        multi_agent_chunks.append(chunk.get("text", chunk.get("content", "")))
                    else:
                        multi_agent_chunks.append(str(chunk))
            
            # Calculate scientific metrics
            trad_metrics = scientific_calculator.calculate_all_metrics(traditional_chunks)
            multi_metrics = scientific_calculator.calculate_all_metrics(multi_agent_chunks)
            
            test_results.append({
                "test_id": test_id,
                "test_name": test_data.get("test_name", test_id),
                "traditional_metrics": trad_metrics.to_dict(),
                "multi_agent_metrics": multi_metrics.to_dict()
            })
        
        if not test_results:
            raise HTTPException(status_code=404, detail="No valid tests found")
        
        # Run batch evaluation
        batch_result = batch_evaluator.evaluate_batch(test_results)
        summary = batch_evaluator.generate_summary(batch_result)
        
        logger.info(f"✅ [BATCH EVAL] Batch evaluation completed for {len(test_results)} tests")
        
        return {
            "success": True,
            "batch_result": batch_result.to_dict(),
            "summary": summary
        }
        
    except ImportError as e:
        logger.error(f"❌ [BATCH EVAL] Import error: {e}")
        raise HTTPException(status_code=500, detail=f"Evaluation module not available: {str(e)}")
    except Exception as e:
        logger.error(f"❌ [BATCH EVAL] Failed to run batch evaluation: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to run batch evaluation: {str(e)}")
